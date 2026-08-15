"""Màn in phiếu giấy báo cáo SX — hỗ trợ nhập hộ khi CN khó dùng máy / mất mạng."""

from __future__ import annotations

import io
from datetime import datetime

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import urlencode

from hrm.permissions import can_view_team_reports, get_team_report_members
from reports.models import DailyWorkReport
from reports.production_hourly import can_proxy_enter_daily_report
from reports.production_shift_policy import shift_display_label
from reports.production_slots import normalize_shift
from reports.report_profile import REPORT_PROFILE_PRODUCTION

User = get_user_model()

DEFAULT_BLANK_ROWS = 10
MAX_BLANK_ROWS = 20


def _parse_prefill_text(request, field: str, *, max_length: int) -> str:
    """Lấy dữ liệu tổ trưởng muốn in sẵn ở dòng đầu của mỗi phiếu."""
    return (request.GET.get(field) or request.POST.get(field) or '').strip()[:max_length]


def _parse_prefill_rows(request, blank_rows: int) -> list[tuple[str, str]]:
    """Lấy mã hàng / công đoạn theo từng dòng từ ô bảng (POST list hoặc 1 giá trị cũ)."""
    codes = request.POST.getlist('row_code') or request.GET.getlist('row_code')
    processes = request.POST.getlist('row_process') or request.GET.getlist('row_process')
    if not codes and not processes:
        single_code = _parse_prefill_text(request, 'product_code', max_length=100)
        single_process = _parse_prefill_text(request, 'process_name', max_length=200)
        codes = [single_code] if single_code else []
        processes = [single_process] if single_process else []
    rows: list[tuple[str, str]] = []
    for index in range(blank_rows):
        code = (codes[index] if index < len(codes) else '')[:100].strip()
        process = (processes[index] if index < len(processes) else '')[:200].strip()
        rows.append((code, process))
    return rows


def _parse_report_date(request):
    raw = (request.GET.get('date') or request.POST.get('date') or '').strip()
    if raw:
        try:
            return datetime.strptime(raw[:10], '%Y-%m-%d').date()
        except ValueError:
            pass
    return timezone.localdate()


def _parse_shift(request) -> str:
    raw = (
        request.GET.get('shift')
        or request.POST.get('shift')
        or DailyWorkReport.SHIFT_MORNING
    ).strip().upper()
    shift = normalize_shift(raw) if raw else DailyWorkReport.SHIFT_MORNING
    if shift not in (DailyWorkReport.SHIFT_MORNING, DailyWorkReport.SHIFT_NIGHT):
        return DailyWorkReport.SHIFT_MORNING
    return shift


def _parse_blank_rows(request) -> int:
    raw = request.GET.get('blank_rows') or request.POST.get('blank_rows')
    try:
        n = int(raw) if raw is not None and str(raw).strip() != '' else DEFAULT_BLANK_ROWS
    except (TypeError, ValueError):
        n = DEFAULT_BLANK_ROWS
    return max(4, min(MAX_BLANK_ROWS, n))


def _parse_selected_user_ids(request) -> list[int]:
    """Nhận for_user / user_ids từ GET hoặc POST (nhiều giá trị hoặc CSV)."""
    raw_list: list[str] = []
    if request.method.upper() == 'POST':
        raw_list.extend(request.POST.getlist('for_user'))
        raw_list.extend(request.POST.getlist('user_ids'))
        csv = (request.POST.get('user_ids_csv') or '').strip()
        if csv:
            raw_list.extend(csv.split(','))
    else:
        raw_list.extend(request.GET.getlist('for_user'))
        raw_list.extend(request.GET.getlist('user_ids'))
        csv = (request.GET.get('user_ids') or '').strip()
        # getlist already covers repeats; also allow comma in a single param
        if csv and ',' in csv and len(request.GET.getlist('user_ids')) == 1:
            raw_list = [p.strip() for p in csv.split(',') if p.strip()]

    ids: list[int] = []
    seen: set[int] = set()
    for raw in raw_list:
        text = str(raw or '').strip()
        if not text:
            continue
        if ',' in text:
            parts = [p.strip() for p in text.split(',') if p.strip()]
        else:
            parts = [text]
        for part in parts:
            try:
                pk = int(part)
            except (TypeError, ValueError):
                continue
            if pk not in seen:
                seen.add(pk)
                ids.append(pk)
    return ids


def _company_context() -> dict:
    try:
        from san_xuat.print_company import (
            COMPANY_ADDRESS,
            COMPANY_NAME,
            COMPANY_TAX_CODE,
        )
        return {
            'company_name': COMPANY_NAME,
            'company_tax_code': COMPANY_TAX_CODE,
            'company_address': COMPANY_ADDRESS,
        }
    except Exception:
        return {
            'company_name': 'JUST PLAY',
            'company_tax_code': '',
            'company_address': '',
        }


def _member_sheet(member, *, report_date, shift: str) -> dict:
    profile = getattr(member, 'profile', None)
    full_name = (
        profile.full_name
        if profile and profile.full_name
        else member.username
    )
    employee_code = (
        profile.employee_code
        if profile and getattr(profile, 'employee_code', None)
        else ''
    )
    department_name = (
        profile.department.name
        if profile and profile.department_id
        else ''
    )
    division_name = (
        profile.division.name
        if profile and getattr(profile, 'division_id', None)
        else ''
    )
    proxy_qs = urlencode({
        'date': report_date.isoformat(),
        'for_user': member.pk,
        'shift': shift,
    })
    return {
        'user_id': member.pk,
        'full_name': full_name,
        'employee_code': employee_code,
        'department_name': department_name,
        'division_name': division_name,
        'proxy_entry_url': f"{reverse('reports:proxy_cn')}?{proxy_qs}",
    }


def _get_print_subjects(request):
    """Trả danh sách CN hợp lệ đã chọn, hoặc None nếu cần quay về chọn lại."""
    team_members = _eligible_team_members(request.user)
    if not team_members:
        messages.info(request, 'Không có công nhân sản xuất cấp dưới để in phiếu.')
        return [], team_members

    by_id = {member.pk: member for member in team_members}
    subjects = []
    for pk in _parse_selected_user_ids(request):
        member = by_id.get(pk)
        if member and can_proxy_enter_daily_report(request.user, member):
            subjects.append(member)
    return subjects, team_members


def _eligible_team_members(viewer):
    return list(
        get_team_report_members(viewer)
        .filter(profile__department__report_profile=REPORT_PROFILE_PRODUCTION)
        .select_related('profile', 'profile__department', 'profile__division')
        .order_by('profile__full_name', 'username')
    )


def _render_select_page(request, *, report_date, shift: str, team_members, blank_rows: int):
    candidates = [
        m for m in team_members
        if can_proxy_enter_daily_report(request.user, m)
    ]
    rows = []
    for member in candidates:
        profile = getattr(member, 'profile', None)
        rows.append({
            'id': member.pk,
            'full_name': (
                profile.full_name
                if profile and profile.full_name
                else member.username
            ),
            'employee_code': (
                profile.employee_code
                if profile and getattr(profile, 'employee_code', None)
                else ''
            ),
            'department_name': (
                profile.department.name
                if profile and profile.department_id
                else ''
            ),
            'division_name': (
                profile.division.name
                if profile and getattr(profile, 'division_id', None)
                else ''
            ),
        })
    back_url = reverse('reports:team_cn') + f'?date={report_date.isoformat()}'
    return render(request, 'reports/print/proxy_paper_select.html', {
        'report_date': report_date,
        'shift': shift,
        'shift_label': shift_display_label(shift),
        'shift_choices': [
            (DailyWorkReport.SHIFT_MORNING, 'Ca sáng'),
            (DailyWorkReport.SHIFT_NIGHT, 'Ca tối'),
        ],
        'members': rows,
        'blank_rows': blank_rows,
        'back_url': back_url,
        'print_url': reverse('reports:proxy_paper_sheet'),
    })


def proxy_paper_sheet(request):
    """Chọn CN → in phiếu ghi sản lượng trống (A4 ngang) + link mở nhập hộ."""
    if not can_view_team_reports(request.user):
        messages.error(request, 'Bạn không có quyền in phiếu / nhập hộ báo cáo.')
        return redirect('home_portal')

    report_date = _parse_report_date(request)
    shift = _parse_shift(request)
    blank_rows = _parse_blank_rows(request)
    selected_ids = _parse_selected_user_ids(request)
    product_code = _parse_prefill_text(request, 'product_code', max_length=100)
    process_name = _parse_prefill_text(request, 'process_name', max_length=200)

    team_members = _eligible_team_members(request.user)
    if not team_members:
        messages.info(request, 'Không có công nhân sản xuất cấp dưới để in phiếu.')
        return redirect('reports:team_cn')

    # Chưa chọn ai → màn chọn (tránh in cả tổ phí giấy).
    # Ngoại lệ: for_user đơn lẻ từ nút từng dòng / nhập hộ → in ngay.
    if not selected_ids:
        return _render_select_page(
            request,
            report_date=report_date,
            shift=shift,
            team_members=team_members,
            blank_rows=blank_rows,
        )

    by_id = {m.pk: m for m in team_members}
    subjects = []
    for pk in selected_ids:
        member = by_id.get(pk)
        if member is None:
            continue
        if not can_proxy_enter_daily_report(request.user, member):
            continue
        subjects.append(member)

    if not subjects:
        messages.error(request, 'Không có công nhân hợp lệ trong danh sách đã chọn.')
        return _render_select_page(
            request,
            report_date=report_date,
            shift=shift,
            team_members=team_members,
            blank_rows=blank_rows,
        )

    sheets = [
        _member_sheet(m, report_date=report_date, shift=shift)
        for m in subjects
    ]

    proxy_entry_url = sheets[0]['proxy_entry_url'] if sheets else ''
    select_qs = urlencode({
        'date': report_date.isoformat(),
        'shift': shift,
        'blank_rows': blank_rows,
    })
    back_url = f"{reverse('reports:proxy_paper_sheet')}?{select_qs}"
    if len(sheets) == 1 and len(selected_ids) == 1:
        # Từ nút in 1 người: quay lại nhập hộ của người đó tiện hơn.
        back_url = sheets[0]['proxy_entry_url']

    ctx = {
        'print_title': f'Phiếu SX {report_date.strftime("%d/%m/%Y")} · {shift_display_label(shift)}',
        'report_date': report_date,
        'shift': shift,
        'shift_label': shift_display_label(shift),
        'sheets': sheets,
        'blank_row_numbers': list(range(1, blank_rows + 1)),
        'blank_rows': blank_rows,
        'proxy_entry_url': proxy_entry_url,
        'back_url': back_url,
        'select_url': f"{reverse('reports:proxy_paper_sheet')}?{select_qs}",
        'excel_url': f"{reverse('reports:proxy_paper_sheet').rstrip('/')}/excel/",
        'product_code': product_code,
        'process_name': process_name,
        'printed_at': timezone.localtime(),
        **_company_context(),
    }
    return render(request, 'reports/print/proxy_paper_sheet.html', ctx)


def proxy_paper_sheet_excel(request):
    """Tải file Excel phiếu giấy cùng định dạng dữ liệu đang chuẩn bị in."""
    if not can_view_team_reports(request.user):
        messages.error(request, 'Bạn không có quyền tải phiếu Excel.')
        return redirect('home_portal')

    report_date = _parse_report_date(request)
    shift = _parse_shift(request)
    blank_rows = _parse_blank_rows(request)
    prefill_rows = _parse_prefill_rows(request, blank_rows)
    subjects, team_members = _get_print_subjects(request)
    if not team_members or not subjects:
        messages.error(request, 'Chọn ít nhất một công nhân hợp lệ để tải phiếu Excel.')
        return _render_select_page(
            request,
            report_date=report_date,
            shift=shift,
            team_members=team_members,
            blank_rows=blank_rows,
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Máy chủ thiếu thư viện tạo file Excel.')
        return redirect('reports:proxy_paper_sheet')

    workbook = Workbook()
    workbook.remove(workbook.active)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(vertical='center', wrap_text=True)
    header_fill = PatternFill('solid', fgColor='FFF200')
    columns = [5, 16, 34, 12, 12, 11, 9, 12, 20]
    headers = [
        'STT', 'Mã hàng', 'Công đoạn - Size', 'Thời gian\nbắt đầu',
        'Thời gian\nkết thúc', 'Số lượng', 'SL lỗi', 'Định mức\n(sp/h)', 'Ghi chú',
    ]

    for index, member in enumerate(subjects, start=1):
        item = _member_sheet(member, report_date=report_date, shift=shift)
        title = (item['full_name'] or f'CN {index}')[:31]
        worksheet = workbook.create_sheet(title=title)
        worksheet.sheet_view.showGridLines = False
        worksheet.page_setup.orientation = 'landscape'
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.3
        worksheet.page_margins.bottom = 0.3
        for col, width in enumerate(columns, start=1):
            worksheet.column_dimensions[get_column_letter(col)].width = width

        worksheet.merge_cells('A1:E1')
        worksheet['A1'] = 'BÁO CÁO SẢN XUẤT HÀNG NGÀY'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A1'].fill = header_fill
        worksheet['A1'].alignment = center
        worksheet.merge_cells('F1:G1')
        worksheet['F1'] = f'Thời gian: {shift_display_label(shift)}'
        worksheet['F1'].font = Font(bold=True)
        worksheet['F1'].alignment = center
        worksheet.merge_cells('H1:I1')
        worksheet['H1'] = f'Ngày: {report_date:%d/%m/%Y}'
        worksheet['H1'].font = Font(bold=True)
        worksheet['H1'].alignment = center
        for cell in worksheet[1]:
            cell.border = border

        metadata = [
            ('Mã NV:', item['employee_code'], 'Họ và Tên:', item['full_name']),
            ('Bộ phận:', item['department_name'], 'Tổ:', item['division_name']),
        ]
        for row_num, values in enumerate(metadata, start=2):
            worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
            worksheet.merge_cells(start_row=row_num, start_column=3, end_row=row_num, end_column=5)
            worksheet.merge_cells(start_row=row_num, start_column=6, end_row=row_num, end_column=7)
            worksheet.merge_cells(start_row=row_num, start_column=8, end_row=row_num, end_column=9)
            for col, value in zip((1, 3, 6, 8), values):
                worksheet.cell(row_num, col, value)
            for cell in worksheet[row_num]:
                cell.border = border
                cell.alignment = left

        header_row = 4
        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(header_row, col, header)
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = border
            cell.fill = PatternFill('solid', fgColor='F3F3F3')
        worksheet.row_dimensions[header_row].height = 32

        for row_offset in range(blank_rows):
            row_num = header_row + 1 + row_offset
            code, process = prefill_rows[row_offset]
            worksheet.cell(row_num, 1, row_offset + 1)
            worksheet.cell(row_num, 2, code)
            worksheet.cell(row_num, 3, process)
            for col in range(1, 10):
                cell = worksheet.cell(row_num, col)
                cell.border = border
                cell.alignment = center if col == 1 else left
            worksheet.row_dimensions[row_num].height = 22

        worksheet.freeze_panes = 'A5'
        last_data_row = header_row + blank_rows
        worksheet.print_area = f'A1:I{last_data_row}'
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1

    output = io.BytesIO()
    workbook.save(output)
    filename = f'phieu_sx_{report_date:%Y%m%d}_{shift.lower()}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

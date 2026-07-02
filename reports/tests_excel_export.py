"""Tests for daily report Excel export."""

from datetime import date
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from hrm.models import Department, DepartmentMenuPermission, Profile, RoleModulePermission
from hrm.permissions import ROLE_EMPLOYEE, ROLE_TEAM_LEADER
from reports.models import DailyWorkReport, ProductionShiftProduct
from reports.production_hourly import (
    ensure_active_work_block,
    ensure_work_day_started,
    finalize_product_with_metadata,
    save_hourly_entry,
)
from reports.report_profile import REPORT_PROFILE_PRODUCTION

User = get_user_model()


class ReportExcelExportTests(TestCase):
    def setUp(self):
        self.dept, _ = Department.objects.get_or_create(
            name='SX Export Test',
            defaults={'report_profile': REPORT_PROFILE_PRODUCTION},
        )
        if self.dept.report_profile != REPORT_PROFILE_PRODUCTION:
            self.dept.report_profile = REPORT_PROFILE_PRODUCTION
            self.dept.save(update_fields=['report_profile'])
        DepartmentMenuPermission.objects.get_or_create(
            department=self.dept,
            defaults={'modules': ['reports']},
        )
        reports_perms = {'reports': {'view': True, 'edit': True}}
        for role in (ROLE_EMPLOYEE, ROLE_TEAM_LEADER):
            RoleModulePermission.objects.update_or_create(
                role=role,
                defaults={'module_permissions': reports_perms},
            )
        self.worker = User.objects.create_user(username='worker_export', password='x')
        Profile.objects.filter(user=self.worker).update(
            full_name='NV Export',
            department=self.dept,
            role=ROLE_EMPLOYEE,
            is_employed=True,
        )
        self.leader = User.objects.create_user(username='leader_export', password='x')
        Profile.objects.filter(user=self.leader).update(
            full_name='TP Export',
            department=self.dept,
            role=ROLE_TEAM_LEADER,
            is_employed=True,
        )
        self.leader.profile.subordinates.add(self.worker)
        self.report_date = date(2026, 6, 18)
        self.report = DailyWorkReport.objects.create(
            employee=self.worker,
            report_date=self.report_date,
            report_profile=REPORT_PROFILE_PRODUCTION,
            shift=DailyWorkReport.SHIFT_MORNING,
            status=DailyWorkReport.STATUS_SUBMITTED,
        )
        self.report.shift = DailyWorkReport.SHIFT_MORNING
        ensure_work_day_started(self.report)
        product = ensure_active_work_block(self.report)
        save_hourly_entry(product, 0, 100)
        finalize_product_with_metadata(
            self.report,
            product_code='PEGASUS',
            process_name='RÁP ĐÁY',
            norm_per_hour=180,
        )

    def test_worker_can_export_own_report(self):
        self.client.force_login(self.worker)
        response = self.client.get(reverse('reports:detail_export_cn', args=[self.report.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml.sheet', response['Content-Type'])
        wb = load_workbook(BytesIO(response.content))
        self.assertIn('Tong_hop', wb.sheetnames)
        self.assertIn('San_luong', wb.sheetnames)

    def test_leader_can_export_subordinate_report(self):
        self.client.force_login(self.leader)
        response = self.client.get(reverse('reports:detail_export_cn', args=[self.report.pk]))
        self.assertEqual(response.status_code, 200)

    def test_history_for_user_lists_subordinate_reports(self):
        self.client.force_login(self.leader)
        response = self.client.get(
            reverse('reports:my_cn'),
            {'for_user': self.worker.pk},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Lịch sử báo cáo — NV Export')
        self.assertContains(response, self.report_date.strftime('%d/%m/%Y'))

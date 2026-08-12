"""VPS smoke test: IE import/export tách theo nhóm / thư viện / routing.

Chạy trên VPS:
  docker compose exec -T web python manage.py shell -c \\
    \"exec(open('scripts/vps_qa_ie_import_export.py', encoding='utf-8').read())\"
"""
from __future__ import annotations

import io
import sys
from datetime import datetime

from django.contrib.auth import get_user_model
from django.test import Client
from django.urls import reverse

from hrm.menu_permissions import user_can_access_menu, user_can_export_menu
from hrm.module_permissions import MODULE_SAN_XUAT, user_can_access_module
from san_xuat.ie_models import SxOperation
from san_xuat.services.operation_master import (
    KIND_GROUPS,
    KIND_LIBRARY,
    KIND_ROUTING,
    LIB_HEADERS,
    SHEET_GUIDE,
    SHEET_GROUP,
    SHEET_LIB,
    SHEET_ROUTING,
    export_ie_dataset_workbook,
    import_ie_dataset,
)

User = get_user_model()
HOST = 'portal.justplay.vn'
PREFIX = '[VPS-QA-IE]'
PASS: list[str] = []
FAIL: list[str] = []
WARN: list[str] = []


def ok(msg: str) -> None:
    PASS.append(msg)
    print(f'  OK  {msg}')


def bad(msg: str) -> None:
    FAIL.append(msg)
    print(f'  FAIL  {msg}')


def warn(msg: str) -> None:
    WARN.append(msg)
    print(f'  WARN  {msg}')


def pick_user():
    """Ưu tiên user có quyền SX + menu ie + export."""
    qs = User.objects.filter(is_active=True).select_related('profile')
    ranked = []
    for u in qs.iterator(chunk_size=200):
        if not user_can_access_module(u, MODULE_SAN_XUAT):
            continue
        score = 0
        if u.is_superuser:
            score += 100
        if user_can_access_menu(u, MODULE_SAN_XUAT, 'ie'):
            score += 20
        if user_can_export_menu(u, MODULE_SAN_XUAT, 'ie'):
            score += 10
        if score:
            ranked.append((score, u.pk, u))
    ranked.sort(reverse=True)
    return ranked[0][2] if ranked else None


def assert_xlsx(resp, label: str) -> bytes | None:
    if resp.status_code != 200:
        bad(f'{label}: HTTP {resp.status_code}')
        return None
    ctype = resp.get('Content-Type', '')
    if 'spreadsheetml' not in ctype and 'octet-stream' not in ctype:
        bad(f'{label}: Content-Type={ctype!r}')
        return None
    body = resp.content or b''
    if len(body) < 100 or body[:2] != b'PK':
        bad(f'{label}: không phải file xlsx (size={len(body)})')
        return None
    ok(f'{label}: xlsx {len(body)} bytes')
    return body


def assert_two_sheets(wb, data_sheet: str, label: str) -> None:
    names = list(wb.sheetnames)
    if SHEET_GUIDE not in names:
        bad(f'{label}: thiếu {SHEET_GUIDE}; sheets={names}')
        return
    if data_sheet not in names:
        bad(f'{label}: thiếu {data_sheet}; sheets={names}')
        return
    if len(names) != 2:
        bad(f'{label}: expect đúng 2 sheet, got {names}')
        return
    ok(f'{label}: sheets={names}')


def main() -> None:
    print(f'{PREFIX} start @ {datetime.now().isoformat(timespec="seconds")}')
    user = pick_user()
    if not user:
        bad('Không tìm thấy user có quyền module san_xuat')
        print(f'{PREFIX} PASS={len(PASS)} FAIL={len(FAIL)} WARN={len(WARN)}')
        print(f'{PREFIX} exit=1')
        return
    print(f'  user={user.username} super={user.is_superuser}')

    for kind, sheet in (
        (KIND_GROUPS, SHEET_GROUP),
        (KIND_LIBRARY, SHEET_LIB),
        (KIND_ROUTING, SHEET_ROUTING),
    ):
        try:
            wb = export_ie_dataset_workbook(kind, template=True)
            assert_two_sheets(wb, sheet, f'template {kind}')
        except Exception as exc:
            bad(f'template {kind}: {exc}')

    c = Client(HTTP_HOST=HOST)
    c.force_login(user)

    pages = (
        ('san_xuat:ie_group_list', 'groups'),
        ('san_xuat:ie_operation_list', 'library'),
        ('san_xuat:ie_routing_list', 'routing'),
    )
    for url_name, kind in pages:
        r = c.get(reverse(url_name))
        if r.status_code != 200:
            bad(f'{url_name} HTTP {r.status_code}')
            continue
        html = r.content.decode('utf-8', errors='replace')
        for needle in ('Import', 'Xuất Excel', 'ieImportModal', 'Tải file mẫu'):
            if needle in html:
                ok(f'{url_name} has {needle!r}')
            else:
                bad(f'{url_name} missing {needle!r}')

        r = c.get(reverse('san_xuat:ie_import_template_kind', kwargs={'kind': kind}))
        sample = assert_xlsx(r, f'mau-excel/{kind}')
        if sample:
            try:
                import openpyxl
                wb2 = openpyxl.load_workbook(io.BytesIO(sample), read_only=True, data_only=True)
                assert_two_sheets(wb2, {
                    'groups': SHEET_GROUP,
                    'library': SHEET_LIB,
                    'routing': SHEET_ROUTING,
                }[kind], f'http template {kind}')
                wb2.close()
            except Exception as exc:
                bad(f'đọc mau {kind}: {exc}')

        r = c.get(reverse('san_xuat:ie_export_kind', kwargs={'kind': kind}))
        exported = assert_xlsx(r, f'xuat-excel/{kind}')
        if exported:
            try:
                import openpyxl
                wb2 = openpyxl.load_workbook(io.BytesIO(exported), read_only=True, data_only=True)
                assert_two_sheets(wb2, {
                    'groups': SHEET_GROUP,
                    'library': SHEET_LIB,
                    'routing': SHEET_ROUTING,
                }[kind], f'http export {kind}')
                wb2.close()
            except Exception as exc:
                bad(f'đọc export {kind}: {exc}')

    # import library dry-run + real (45s → 0.75 SMV)
    stamp = datetime.now().strftime('%H%M%S')
    op_code = f'QA-IE-{stamp}'
    try:
        from openpyxl import Workbook
        from san_xuat.services.operation_master import _write_sheet

        wb_in = Workbook()
        ws = wb_in.active
        ws.title = SHEET_GUIDE
        ws.append(['QA'])
        ws2 = wb_in.create_sheet(SHEET_LIB)
        row = [''] * len(LIB_HEADERS)
        mapping = {
            'MÃ NHÓM': 'QA',
            'MÃ CÔNG ĐOẠN': op_code,
            'PHIÊN BẢN': 'R01',
            'TÊN CÔNG ĐOẠN': f'QA import {stamp}',
            'ĐỊNH MỨC THỜI GIAN': 45,
            'TRẠNG THÁI': 'Nháp',
            'NOTES': PREFIX,
        }
        for i, h in enumerate(LIB_HEADERS):
            if h in mapping:
                row[i] = mapping[h]
        _write_sheet(ws2, LIB_HEADERS, [row])
        buf = io.BytesIO()
        wb_in.save(buf)
        buf.seek(0)

        dry = import_ie_dataset(buf, KIND_LIBRARY, dry_run=True, user=user)
        ok(f'dry-run created={dry.total_created} updated={dry.total_updated} warns={len(dry.warnings)}')
        if SxOperation.objects.filter(op_code=op_code, op_rev='R01').exists():
            bad('dry-run vẫn ghi DB')
        else:
            ok('dry-run không ghi DB')

        buf.seek(0)
        real = import_ie_dataset(buf, KIND_LIBRARY, dry_run=False, user=user)
        ok(f'import created={real.total_created} updated={real.total_updated}')
        op = SxOperation.objects.filter(op_code=op_code, op_rev='R01').first()
        if not op:
            bad(f'không thấy OP {op_code}/R01 sau import')
        else:
            smv = float(op.base_smv_min or 0)
            if abs(smv - 0.75) < 0.0001:
                ok(f'SMV đúng 0.75 phút (từ 45 giây) op_id={op.pk}')
            else:
                bad(f'SMV sai: {smv} (expect 0.75)')
            label = f'{op.op_code}/{op.op_rev}'
            op.delete()
            ok(f'cleanup deleted {label}')
    except Exception as exc:
        bad(f'import flow: {exc}')
        import traceback
        traceback.print_exc()

    try:
        from django.core.files.uploadedfile import SimpleUploadedFile

        wb_http = export_ie_dataset_workbook(KIND_LIBRARY, template=True)
        raw = io.BytesIO()
        wb_http.save(raw)
        upload = SimpleUploadedFile(
            'mau_qa.xlsx',
            raw.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        r = c.post(
            reverse('san_xuat:ie_operation_list'),
            data={'action': 'import', 'ie_kind': KIND_LIBRARY, 'dry_run': '1', 'excel_file': upload},
        )
        if r.status_code in (302, 200):
            ok(f'POST import dry-run HTTP {r.status_code}')
        else:
            bad(f'POST import dry-run HTTP {r.status_code}')
    except Exception as exc:
        bad(f'POST import: {exc}')

    print()
    print(f'{PREFIX} PASS={len(PASS)} FAIL={len(FAIL)} WARN={len(WARN)}')
    for m in FAIL:
        print(f'  - {m}')
    print(f'{PREFIX} exit={0 if not FAIL else 1}')


main()

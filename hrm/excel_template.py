"""File Excel mẫu import nhân sự — nhiều sheet, danh mục cơ cấu."""

from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from hrm.choices import EXCEL_ALL_HEADERS
from hrm.models import Department, Division, DivisionPosition


def _sample_import_rows() -> list[dict]:
    return [
        {
            'Mã NS': 'NV001',
            'Họ và tên': 'Nguyễn Văn An',
            'Account': 'Annt',
            'Phòng ban': 'SẢN XUẤT',
            'Bộ phận': 'QC',
            'Vị trí': 'Công nhân may',
            'Chức vụ': 'Nhân viên',
            'Ngày vào': '01/01/2026',
            'Ngày sinh': '15/05/1995',
            'Giới tính': 'Nam',
            'Vai trò HT': 'Nhân viên',
            'Nhóm quyền': '',
            'Trạng thái': 'Đang làm',
            'password': 'JustPlay@123',
            'email': 'annt@justplay.vn',
        },
        {
            'Mã NS': 'NV002',
            'Họ và tên': 'Trần Thị Bình',
            'Account': '',
            'Phòng ban': 'SẢN XUẤT',
            'Bộ phận': 'QC',
            'Vị trí': 'Nhân viên QC',
            'Chức vụ': 'Nhân viên',
            'Ngày vào': '10/02/2026',
            'Ngày sinh': '20/08/1998',
            'Giới tính': 'Nữ',
            'Vai trò HT': 'Tổ trưởng',
            'Nhóm quyền': '',
            'Trạng thái': 'Đang làm',
            'password': '',
            'email': '',
        },
        {
            'Mã NS': '',
            'Họ và tên': 'Lê Văn Cường',
            'Account': '',
            'Phòng ban': 'ĐẢM BẢO CHẤT LƯỢNG',
            'Bộ phận': 'QA',
            'Vị trí': 'Trưởng ca',
            'Chức vụ': 'Tổ trưởng',
            'Ngày vào': '',
            'Ngày sinh': '',
            'Giới tính': 'Nam',
            'Vai trò HT': 'Trưởng phòng',
            'Nhóm quyền': '',
            'Trạng thái': 'Đang làm',
            'password': '',
            'email': '',
        },
    ]


def _catalog_rows() -> tuple[list[dict], list[dict], list[dict]]:
    dept_rows = []
    div_rows = []
    pos_rows = []
    from django.db.models import Prefetch

    pos_qs = DivisionPosition.objects.filter(is_active=True).order_by('sort_order', 'name')
    div_qs = Division.objects.filter(is_active=True).prefetch_related(
        Prefetch('positions', queryset=pos_qs),
    ).order_by('sort_order', 'name')
    departments = Department.objects.filter(is_active=True).prefetch_related(
        Prefetch('divisions', queryset=div_qs),
    ).order_by('sort_order', 'name')
    for dept in departments:
        dept_rows.append({'Phòng ban': dept.name, 'Thứ tự': dept.sort_order})
        for div in dept.divisions.all():
            div_rows.append({
                'Phòng ban': dept.name,
                'Bộ phận': div.name,
                'Thứ tự': div.sort_order,
            })
            for pos in div.positions.filter(is_active=True).order_by('sort_order', 'name'):
                pos_rows.append({
                    'Phòng ban': dept.name,
                    'Bộ phận': div.name,
                    'Vị trí (danh mục)': pos.name,
                    'Thứ tự': pos.sort_order,
                })
    unassigned = Division.objects.filter(department__isnull=True, is_active=True).order_by('sort_order', 'name')
    for div in unassigned:
        div_rows.append({'Phòng ban': '(Chưa gán PB)', 'Bộ phận': div.name, 'Thứ tự': div.sort_order})
    return dept_rows, div_rows, pos_rows


def build_import_template_xlsx() -> bytes:
    """Workbook: Nhập liệu + Danh mục PB/BP/Vị trí + Hướng dẫn."""
    output = io.BytesIO()
    sample_df = pd.DataFrame(_sample_import_rows(), columns=EXCEL_ALL_HEADERS)
    dept_rows, div_rows, pos_rows = _catalog_rows()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sample_df.to_excel(writer, index=False, sheet_name='Nhap_lieu')
        pd.DataFrame(dept_rows or [{'Phòng ban': '(Chưa có)', 'Thứ tự': 0}]).to_excel(
            writer, index=False, sheet_name='Phong_ban',
        )
        pd.DataFrame(div_rows or [{'Phòng ban': '', 'Bộ phận': '', 'Thứ tự': 0}]).to_excel(
            writer, index=False, sheet_name='Bo_phan',
        )
        pd.DataFrame(pos_rows or [{'Phòng ban': '', 'Bộ phận': '', 'Vị trí (danh mục)': '', 'Thứ tự': 0}]).to_excel(
            writer, index=False, sheet_name='Vi_tri',
        )
        guide = pd.DataFrame([
            {'Mục': 'Bắt buộc', 'Nội dung': 'Cột Họ và tên — không để trống.'},
            {'Mục': 'Account', 'Nội dung': 'Để trống → hệ thống tự sinh khi thêm mới.'},
            {'Mục': 'Phòng ban / Bộ phận', 'Nội dung': 'Phải khớp sheet Phong_ban, Bo_phan (đúng chính tả).'},
            {'Mục': 'Vị trí', 'Nội dung': 'Khớp sheet Vi_tri hoặc tên trên hồ sơ; import sẽ tạo danh mục vị trí nếu thiếu.'},
            {'Mục': 'Ngày', 'Nội dung': 'Định dạng dd/mm/yyyy (vd: 01/03/2026).'},
            {'Mục': 'Giới tính', 'Nội dung': 'Nam hoặc Nữ.'},
            {'Mục': 'Vai trò HT', 'Nội dung': 'Nhân viên, Tổ trưởng, Trưởng bộ phận, Trưởng phòng, Giám đốc — hoặc mã EMPLOYEE, TEAM_LEADER, …'},
            {'Mục': 'Nhóm quyền', 'Nội dung': 'Tên hoặc mã slug nhóm quyền; để trống → nhóm mặc định theo vai trò.'},
            {'Mục': 'Trạng thái', 'Nội dung': 'Đang làm hoặc Nghỉ việc.'},
            {'Mục': 'Trùng Account/Mã NS', 'Nội dung': 'Cập nhật nhân sự có sẵn, không tạo trùng.'},
            {'Mục': 'Sơ đồ tổ chức', 'Nội dung': 'Sau import, mở Cơ cấu tổ chức → bấm Vị trí để xem danh sách NV.'},
        ])
        guide.to_excel(writer, index=False, sheet_name='Huong_dan')

    output.seek(0)
    wb = load_workbook(output)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0F766E')
    for sheet_name in ('Nhap_lieu', 'Phong_ban', 'Bo_phan', 'Vi_tri', 'Huong_dan'):
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.freeze_panes = 'A2'
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 12
            for row in range(1, min(ws.max_row, 80) + 1):
                val = ws[f'{letter}{row}'].value
                if val is not None:
                    max_len = max(max_len, min(len(str(val)) + 2, 48))
            ws.column_dimensions[letter].width = max_len
    wb.active = wb['Nhap_lieu']
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

from django.contrib.auth.models import User
from django.test import Client, TestCase
from django.urls import reverse

from PortalJustPlay.utils import (
    build_hm_username_base,
    generate_hm_email,
    generate_hm_username,
    next_employee_code,
)

from hrm.models import (
    Department,
    DepartmentMenuPermission,
    Division,
    PermissionGroup,
    Profile,
    RoleModulePermission,
)
from hrm.module_permissions import (
    ALL_MODULE_KEYS,
    MODULE_CHOICES,
    MODULE_HRM,
    MODULE_KPI,
    MODULE_PERMISSIONS,
    MODULE_RECRUITMENT,
    MODULE_TRAINING,
    can_manage_permissions,
    get_department_enabled_modules,
    resolve_module_from_request,
    user_can_access_module,
    user_can_edit_module,
    user_can_export_module,
)
from hrm.group_permissions import normalize_group_permissions, permissions_from_legacy_role
from hrm.permissions import ROLE_DIRECTOR, ROLE_EMPLOYEE, ROLE_TEAM_LEADER


class HMUsernameSuggestionTests(TestCase):
    def test_build_username_examples(self):
        self.assertEqual(build_hm_username_base('Nguyễn Thành Nam'), 'nam.nt')
        self.assertEqual(build_hm_username_base('Trần Thái Viết Hưng'), 'hung.ttv')

    def test_generate_username_avoids_duplicate_with_suffix(self):
        User.objects.create_user(username='nam.nt', password='x')
        self.assertEqual(generate_hm_username('Nguyễn Thành Nam'), 'nam.nt1')

    def test_generate_username_ignores_self_on_edit(self):
        user = User.objects.create_user(username='nam.nt', password='x')
        self.assertEqual(
            generate_hm_username('Nguyễn Thành Nam', exclude_user_id=user.pk),
            'nam.nt',
        )

    def test_generate_email_from_username(self):
        self.assertEqual(generate_hm_email('nam.nt'), 'nam.nt@justplay.vn')

    def test_next_employee_code_increments_from_max(self):
        user = User.objects.create_user(username='code1', password='x')
        Profile.objects.filter(user=user).update(employee_code='440')
        self.assertEqual(next_employee_code(), '441')


from hrm.role_permissions import get_role_permissions, role_allows_edit, role_allows_view


class PermissionLogicTests(TestCase):
    def setUp(self):
        self.dept_hr = Department.objects.create(name='Phòng HR Test', sort_order=1)
        self.dept_xuong = Department.objects.create(name='Phòng Xưởng Test', sort_order=2)

        DepartmentMenuPermission.objects.create(
            department=self.dept_hr,
            modules=['announcements', 'hrm', 'kpi', 'training', 'documents', 'permissions'],
        )
        DepartmentMenuPermission.objects.create(
            department=self.dept_xuong,
            modules=['announcements', 'training', 'assessment'],
        )
        # dept không cấu hình module → full quyền (list rỗng)
        self.dept_full = Department.objects.create(name='Phòng Full Test', sort_order=3)
        DepartmentMenuPermission.objects.create(department=self.dept_full, modules=[])

        RoleModulePermission.objects.update_or_create(
            role=ROLE_EMPLOYEE,
            defaults={
                'module_permissions': {
                    'announcements': {'view': True, 'edit': False},
                    'training': {'view': True, 'edit': False},
                    'hrm': {'view': False, 'edit': False},
                    'kpi': {'view': True, 'edit': False},
                    'recruitment': {'view': False, 'edit': False},
                },
            },
        )
        RoleModulePermission.objects.update_or_create(
            role=ROLE_TEAM_LEADER,
            defaults={
                'module_permissions': {
                    'announcements': {'view': True, 'edit': False},
                    'training': {'view': True, 'edit': False},
                    'hrm': {'view': True, 'edit': False},
                    'kpi': {'view': True, 'edit': True},
                    'reports': {'view': True, 'edit': True},
                },
            },
        )

        self.employee = User.objects.create_user(
            username='perm_employee',
            password='testpass123',
        )
        Profile.objects.filter(user=self.employee).update(
            department=self.dept_xuong,
            role=ROLE_EMPLOYEE,
            full_name='NV Test',
            permission_group=None,
        )
        self.employee.refresh_from_db()

        self.team_leader = User.objects.create_user(
            username='perm_leader',
            password='testpass123',
        )
        Profile.objects.filter(user=self.team_leader).update(
            department=self.dept_hr,
            role=ROLE_TEAM_LEADER,
            full_name='Tổ trưởng Test',
            permission_group=None,
        )
        self.team_leader.refresh_from_db()

        self.hr_editor = User.objects.create_user(
            username='perm_hr_editor',
            password='testpass123',
            is_staff=True,
        )
        Profile.objects.filter(user=self.hr_editor).update(
            department=self.dept_hr,
            role=ROLE_DIRECTOR,
            full_name='HR Director Test',
            permission_group=None,
        )
        self.hr_editor.refresh_from_db()

    def test_empty_department_modules_means_full_access(self):
        enabled = get_department_enabled_modules(self.dept_full)
        self.assertEqual(len(enabled), len(ALL_MODULE_KEYS))
        self.assertIn(MODULE_RECRUITMENT, enabled)

    def test_department_restricts_modules(self):
        enabled = get_department_enabled_modules(self.dept_xuong)
        self.assertIn(MODULE_TRAINING, enabled)
        self.assertNotIn(MODULE_HRM, enabled)

    def test_employee_cannot_view_hrm_even_if_dept_allows(self):
        Profile.objects.filter(user=self.team_leader).update(role=ROLE_EMPLOYEE)
        self.team_leader.refresh_from_db()
        self.assertFalse(user_can_access_module(self.team_leader, MODULE_HRM))

    def test_team_leader_can_view_not_edit_hrm_when_role_configured(self):
        self.assertTrue(user_can_access_module(self.team_leader, MODULE_HRM))
        self.assertFalse(user_can_edit_module(self.team_leader, MODULE_HRM))

    def test_kpi_hidden_from_portal_permissions(self):
        self.assertFalse(user_can_edit_module(self.team_leader, MODULE_KPI))
        self.assertFalse(user_can_access_module(self.team_leader, MODULE_KPI))

    def test_employee_can_view_training(self):
        self.assertTrue(user_can_access_module(self.employee, MODULE_TRAINING))

    def test_director_in_hr_dept_can_edit_hrm(self):
        self.assertTrue(user_can_edit_module(self.hr_editor, MODULE_HRM))

    def test_resolve_module_urls(self):
        self.assertEqual(resolve_module_from_request('/dashboard/users/'), MODULE_HRM)
        self.assertEqual(resolve_module_from_request('/tai-lieu/'), 'documents')
        self.assertEqual(resolve_module_from_request('/dashboard/permissions/'), MODULE_PERMISSIONS)
        self.assertEqual(
            resolve_module_from_request('/dashboard/departments/1/permissions/'),
            MODULE_PERMISSIONS,
        )
        self.assertEqual(resolve_module_from_request('/dashboard/permissions/roles/EMPLOYEE/'), MODULE_PERMISSIONS)
        self.assertEqual(
            resolve_module_from_request('/dashboard/', 'recruitment'),
            MODULE_RECRUITMENT,
        )

    def test_can_manage_permissions_requires_permissions_edit(self):
        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={
                'module_permissions': {
                    'permissions': {'view': True, 'edit': True},
                    'hrm': {'view': True, 'edit': True},
                },
            },
        )
        self.assertTrue(can_manage_permissions(self.hr_editor))
        self.assertFalse(can_manage_permissions(self.employee))

    def test_hrm_edit_without_permissions_module_cannot_manage(self):
        """Có quyền HRM nhưng không có module Phân quyền — không vào cấu hình."""
        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={
                'module_permissions': {
                    'hrm': {'view': True, 'edit': True},
                    'permissions': {'view': False, 'edit': False},
                },
            },
        )
        Profile.objects.filter(user=self.hr_editor).update(role=ROLE_DIRECTOR)
        self.hr_editor.refresh_from_db()
        self.assertFalse(can_manage_permissions(self.hr_editor))

    def test_role_permission_defaults_seeded(self):
        perms = get_role_permissions(ROLE_DIRECTOR)
        self.assertTrue(perms[MODULE_HRM]['edit'])


class PermissionGroupTests(TestCase):
    def setUp(self):
        self.dept_hr = Department.objects.create(name='PG HR', sort_order=1)
        DepartmentMenuPermission.objects.create(
            department=self.dept_hr,
            modules=['hrm', 'announcements', 'training'],
        )

        employee_legacy = permissions_from_legacy_role(ROLE_EMPLOYEE)
        self.group_sx = PermissionGroup.objects.create(
            slug='test-nv-sx',
            name='Test NV Sản xuất',
            module_permissions=employee_legacy,
        )
        hcns_perms = normalize_group_permissions(employee_legacy)
        hcns_perms['hrm'] = {
            'view': True,
            'create': True,
            'update': True,
            'delete': False,
            'export': True,
        }
        self.group_hcns = PermissionGroup.objects.create(
            slug='test-nv-hcns',
            name='Test NV HCNS',
            module_permissions=hcns_perms,
        )

        self.sx_user = User.objects.create_user(username='pg_sx', password='testpass123')
        Profile.objects.filter(user=self.sx_user).update(
            department=self.dept_hr,
            role=ROLE_EMPLOYEE,
            permission_group=self.group_sx,
        )
        self.hcns_user = User.objects.create_user(username='pg_hcns', password='testpass123')
        Profile.objects.filter(user=self.hcns_user).update(
            department=self.dept_hr,
            role=ROLE_EMPLOYEE,
            permission_group=self.group_hcns,
        )
        self.hcns_user.refresh_from_db()

    def test_production_staff_cannot_edit_hrm(self):
        self.assertTrue(user_can_access_module(self.sx_user, MODULE_HRM) is False or not user_can_edit_module(self.sx_user, MODULE_HRM))
        self.assertFalse(user_can_edit_module(self.sx_user, MODULE_HRM))

    def test_hcns_staff_can_edit_and_export_hrm(self):
        self.assertTrue(user_can_edit_module(self.hcns_user, MODULE_HRM))
        self.assertTrue(user_can_export_module(self.hcns_user, MODULE_HRM))

    def test_hcns_staff_granular_hrm_permissions(self):
        from hrm.module_permissions import (
            user_can_create_module,
            user_can_delete_module,
            user_can_update_module,
        )

        hcns_perms = normalize_group_permissions(permissions_from_legacy_role(ROLE_EMPLOYEE))
        hcns_perms['hrm'] = {
            'view': True,
            'create': True,
            'update': True,
            'delete': False,
            'export': True,
        }
        self.group_hcns.module_permissions = hcns_perms
        self.group_hcns.save(update_fields=['module_permissions'])
        self.hcns_user.refresh_from_db()

        self.assertTrue(user_can_create_module(self.hcns_user, MODULE_HRM))
        self.assertTrue(user_can_update_module(self.hcns_user, MODULE_HRM))
        self.assertFalse(user_can_delete_module(self.hcns_user, MODULE_HRM))

    def test_permission_group_urls_resolve(self):
        self.assertEqual(
            resolve_module_from_request('/dashboard/permissions/groups/1/edit/'),
            MODULE_PERMISSIONS,
        )

    def test_permission_form_clears_export_for_modules_without_excel(self):
        from hrm.forms import PermissionGroupPermissionForm
        from hrm.group_permissions import MODULE_SUPPORTS_EXPORT

        data = {f'export_{key}': 'on' for key, _ in MODULE_CHOICES}
        data.update({f'view_{key}': 'on' for key, _ in MODULE_CHOICES})
        form = PermissionGroupPermissionForm(data)
        self.assertTrue(form.is_valid())
        perms = form.cleaned_permissions()
        for key, _ in MODULE_CHOICES:
            if key in MODULE_SUPPORTS_EXPORT:
                self.assertTrue(perms[key]['export'], key)
            else:
                self.assertFalse(perms[key]['export'], key)

    def test_permission_form_clears_create_update_delete_for_audit(self):
        from hrm.forms import PermissionGroupPermissionForm

        data = {f'{action}_audit': 'on' for action in ('view', 'create', 'update', 'delete', 'export')}
        form = PermissionGroupPermissionForm(data)
        self.assertTrue(form.is_valid())
        perms = form.cleaned_permissions()['audit']
        self.assertTrue(perms['view'])
        self.assertTrue(perms['export'])
        self.assertFalse(perms['create'])
        self.assertFalse(perms['update'])
        self.assertFalse(perms['delete'])

    def test_module_choices_aligned_with_menu_and_group_matrix(self):
        from hrm.forms import PERM_GROUP_MODULE_ICONS, PermissionGroupPermissionForm
        from hrm.group_permissions import MODULE_LIST_META
        from hrm.module_permissions import DEPARTMENT_MENU_SECTIONS

        keys = {key for key, _ in MODULE_CHOICES}
        self.assertEqual(keys, ALL_MODULE_KEYS)
        self.assertEqual(set(MODULE_LIST_META), keys)
        self.assertEqual(set(PERM_GROUP_MODULE_ICONS), keys)

        form = PermissionGroupPermissionForm()
        matrix_keys = {row['key'] for row in form.module_rows()}
        self.assertEqual(matrix_keys, keys)

        menu_keys = set()
        for section in DEPARTMENT_MENU_SECTIONS:
            menu_keys.update(section['modules'])
        self.assertEqual(menu_keys, keys)

    def test_department_menu_form_groups_match_sidebar_labels(self):
        from hrm.forms import DepartmentMenuPermissionForm
        from hrm.module_permissions import DEPARTMENT_MENU_SECTIONS, MODULE_LABELS

        form = DepartmentMenuPermissionForm(initial={'modules': sorted(ALL_MODULE_KEYS)})
        sections = form.menu_section_rows()
        self.assertEqual(len(sections), len(DEPARTMENT_MENU_SECTIONS))
        row_labels = [row['label'] for section in sections for row in section['rows']]
        self.assertIn(MODULE_LABELS['documents'], row_labels)
        self.assertIn(MODULE_LABELS['audit'], row_labels)
        self.assertEqual(
            {row['key'] for section in sections for row in section['rows']},
            ALL_MODULE_KEYS,
        )


class HrmGranularPermissionViewTests(TestCase):
    def setUp(self):
        self.dept_hr = Department.objects.create(name='HCNS Granular', sort_order=1)
        DepartmentMenuPermission.objects.create(
            department=self.dept_hr,
            modules=['hrm'],
        )
        employee_legacy = permissions_from_legacy_role(ROLE_EMPLOYEE)
        hcns_perms = normalize_group_permissions(employee_legacy)
        hcns_perms['hrm'] = {
            'view': True,
            'create': True,
            'update': True,
            'delete': False,
            'export': True,
        }
        self.group = PermissionGroup.objects.create(
            slug='test-hcns-granular',
            name='Test HCNS granular',
            module_permissions=hcns_perms,
        )
        self.hr_user = User.objects.create_user(username='hcns_granular', password='testpass123')
        Profile.objects.filter(user=self.hr_user).update(
            department=self.dept_hr,
            role=ROLE_EMPLOYEE,
            permission_group=self.group,
        )
        self.target = User.objects.create_user(username='victim_user', password='testpass123')
        Profile.objects.filter(user=self.target).update(
            department=self.dept_hr,
            role=ROLE_EMPLOYEE,
            full_name='Victim User',
        )
        self.client = Client(HTTP_HOST='testserver')
        self.client.force_login(self.hr_user)

    def test_hcns_can_list_add_edit_but_not_delete(self):
        self.assertEqual(self.client.get(reverse('user_list')).status_code, 200)
        self.assertEqual(self.client.get(reverse('user_add')).status_code, 200)
        self.assertEqual(self.client.get(reverse('user_edit', args=[self.target.id])).status_code, 200)
        response = self.client.get(reverse('user_delete', args=[self.target.id]))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('home_portal'))
        self.assertTrue(User.objects.filter(username='victim_user').exists())

    def test_hcns_user_list_hides_delete_action(self):
        response = self.client.get(reverse('user_list'))
        self.assertNotContains(response, reverse('user_delete', args=[self.target.id]))
        self.assertContains(response, reverse('user_edit', args=[self.target.id]))


class HrmViewOnlyUserEditTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name='View Only Dept', sort_order=2)
        DepartmentMenuPermission.objects.create(department=self.dept, modules=['hrm'])
        perms = normalize_group_permissions(permissions_from_legacy_role(ROLE_EMPLOYEE))
        perms['hrm'] = {
            'view': True,
            'create': False,
            'update': False,
            'delete': False,
            'export': False,
        }
        self.group = PermissionGroup.objects.create(
            slug='test-hrm-view-only',
            name='HRM view only',
            module_permissions=perms,
        )
        self.viewer = User.objects.create_user(username='hrm_viewer', password='testpass123')
        Profile.objects.filter(user=self.viewer).update(
            department=self.dept,
            role=ROLE_EMPLOYEE,
            permission_group=self.group,
            is_employed=True,
        )
        self.target = User.objects.create_user(username='view_target', password='testpass123')
        Profile.objects.filter(user=self.target).update(
            department=self.dept,
            role=ROLE_EMPLOYEE,
            full_name='View Target',
            is_employed=True,
        )
        self.client = Client(HTTP_HOST='testserver')
        self.client.force_login(self.viewer)

    def test_view_only_can_open_user_edit_readonly(self):
        response = self.client.get(reverse('user_edit', args=[self.target.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'jp-user-form--view-mode')
        self.assertContains(response, 'Chỉ xem')
        self.assertNotContains(response, 'class="btn btn-sm btn-outline-hm fw-bold jp-user-tab-edit-btn"')

    def test_view_only_cannot_post_user_edit(self):
        response = self.client.post(reverse('user_edit', args=[self.target.id]), {
            'username': 'view_target',
            'full_name': 'Changed Name',
            'role': ROLE_EMPLOYEE,
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('home_portal'))


class DepartmentPermissionTemplateTests(TestCase):
    def test_each_department_has_two_groups(self):
        from hrm.department_permission_templates import DEPARTMENT_PERMISSION_TEMPLATES
        self.assertEqual(len(DEPARTMENT_PERMISSION_TEMPLATES), 9)
        for item in DEPARTMENT_PERMISSION_TEMPLATES:
            self.assertIn('employee', item)
            self.assertIn('manager', item)

    def test_hcns_employee_can_edit_hrm_in_matrix(self):
        from hrm.department_permission_templates import DEPARTMENT_PERMISSION_TEMPLATES
        hcns = next(t for t in DEPARTMENT_PERMISSION_TEMPLATES if t['code'] == 'hcns')
        hrm = hcns['employee']['hrm']
        self.assertTrue(hrm['view'])
        self.assertTrue(hrm['create'])
        self.assertTrue(hrm['update'])
        self.assertFalse(hrm['delete'])


class PermissionMiddlewareTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name='MW Dept', sort_order=1)
        DepartmentMenuPermission.objects.create(
            department=self.dept,
            modules=['announcements'],
        )
        RoleModulePermission.objects.update_or_create(
            role=ROLE_EMPLOYEE,
            defaults={
                'module_permissions': {
                    'announcements': {'view': True, 'edit': False},
                    'kpi': {'view': True, 'edit': False},
                },
            },
        )
        self.user = User.objects.create_user(username='mw_user', password='testpass123')
        Profile.objects.filter(user=self.user).update(
            department=self.dept,
            role=ROLE_EMPLOYEE,
            permission_group=None,
        )

    def test_middleware_blocks_kpi_url(self):
        client = Client(HTTP_HOST='testserver')
        client.force_login(self.user)
        response = client.get('/kpi/', follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('home_portal'))

    def test_middleware_allows_announcements(self):
        client = Client(HTTP_HOST='testserver')
        client.force_login(self.user)
        response = client.get('/announcements/')
        self.assertEqual(response.status_code, 200)

    def test_agent_gate_removed_from_portal(self):
        """Trang cài/đăng ký agent đã gỡ — không còn redirect loop."""
        client = Client(HTTP_HOST='testserver')
        client.force_login(self.user)
        response = client.get('/', follow=False)
        self.assertEqual(response.status_code, 200)

    def test_middleware_blocks_hrm_user_list_for_employee(self):
        """Nhân viên xưởng không có quyền HRM — chặn /dashboard/users/."""
        dept = Department.objects.create(name='MW2', sort_order=9)
        DepartmentMenuPermission.objects.create(
            department=dept,
            modules=['announcements', 'hrm'],
        )
        user = User.objects.create_user(username='mw_hrm_block', password='testpass123')
        Profile.objects.filter(user=user).update(department=dept, role=ROLE_EMPLOYEE)
        client = Client(HTTP_HOST='testserver')
        client.force_login(user)
        response = client.get('/dashboard/users/', follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('home_portal'))

    def test_middleware_blocks_permission_config_without_module(self):
        dept = Department.objects.create(name='MW Perm', sort_order=10)
        DepartmentMenuPermission.objects.create(
            department=dept,
            modules=['announcements'],
        )
        user = User.objects.create_user(username='mw_perm', password='testpass123')
        Profile.objects.filter(user=user).update(department=dept, role=ROLE_EMPLOYEE)
        client = Client(HTTP_HOST='testserver')
        client.force_login(user)
        response = client.get('/dashboard/permissions/', follow=False)
        self.assertEqual(response.status_code, 302)

    def test_admin_only_blocks_employee_from_user_list(self):
        client = Client(HTTP_HOST='testserver')
        client.force_login(self.user)
        response = client.get('/dashboard/users/')
        self.assertEqual(response.status_code, 302)


class RolePermissionFormTests(TestCase):
    def test_normalize_edit_implies_view(self):
        from hrm.forms import RolePermissionForm

        form = RolePermissionForm(data={
            'view_hrm': False,
            'edit_hrm': True,
            'view_announcements': True,
            'edit_announcements': False,
        })
        self.assertTrue(form.is_valid(), form.errors)
        perms = form.cleaned_permissions()
        self.assertTrue(perms[MODULE_HRM]['view'])
        self.assertTrue(perms[MODULE_HRM]['edit'])


class ProfileAvatarTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='avatar_user', password='testpass123')
        Profile.objects.filter(user=self.user).update(full_name='Avatar Test')
        self.client = Client()
        self.client.force_login(self.user)

    @staticmethod
    def _sample_jpeg_bytes(width=320, height=240):
        from io import BytesIO

        from PIL import Image

        buf = BytesIO()
        Image.new('RGB', (width, height), color=(220, 38, 38)).save(buf, format='JPEG')
        return buf.getvalue()

    def test_prepare_avatar_image_resizes_to_150(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from hrm.avatar_utils import AVATAR_SIZE, prepare_avatar_image

        upload = SimpleUploadedFile('wide.jpg', self._sample_jpeg_bytes(), content_type='image/jpeg')
        prepared = prepare_avatar_image(upload)
        from PIL import Image

        img = Image.open(prepared)
        self.assertEqual(img.size, (AVATAR_SIZE, AVATAR_SIZE))

    def test_update_avatar(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        image = SimpleUploadedFile('me.jpg', self._sample_jpeg_bytes(), content_type='image/jpeg')
        response = self.client.post(reverse('update_avatar'), {
            'avatar': image,
            'next': '/',
        })
        self.assertEqual(response.status_code, 302)
        self.user.profile.refresh_from_db()
        self.assertTrue(self.user.profile.avatar)
        self.assertIn('me', self.user.profile.avatar.name)

    def test_user_list_avatar_zoomable(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        image = SimpleUploadedFile('face.jpg', self._sample_jpeg_bytes(), content_type='image/jpeg')
        self.client.post(reverse('update_avatar'), {'avatar': image, 'next': '/'})
        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={'module_permissions': {MODULE_HRM: {'view': True, 'edit': True}}},
        )
        Profile.objects.filter(user=self.user).update(role=ROLE_DIRECTOR, is_employed=True, permission_group=None)
        response = self.client.get(reverse('user_list'))
        self.assertContains(response, 'data-jp-avatar-zoom')
        self.assertContains(response, 'jpAvatarZoomModal')


class UserDeleteTests(TestCase):
    def setUp(self):
        self.dept = Department.objects.create(name='HR Delete Dept', sort_order=1)
        DepartmentMenuPermission.objects.create(department=self.dept, modules=['hrm'])
        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={
                'module_permissions': {
                    MODULE_HRM: {'view': True, 'edit': True},
                },
            },
        )
        self.admin = User.objects.create_user(username='hr_del', password='testpass123', is_staff=True)
        Profile.objects.filter(user=self.admin).update(
            department=self.dept,
            role=ROLE_DIRECTOR,
            full_name='HR Delete',
            is_employed=True,
        )
        self.client = Client(HTTP_HOST='testserver')
        self.client.force_login(self.admin)

    def test_cannot_delete_system_admin_account(self):
        system_admin = User.objects.create_user(username='admin', password='testpass123', is_superuser=True)
        response = self.client.get(reverse('user_delete', args=[system_admin.id]))
        self.assertEqual(response.status_code, 302)
        self.assertTrue(User.objects.filter(username='admin').exists())

    def test_can_delete_non_admin_superuser(self):
        loadtest = User.objects.create_user(
            username='loadtest',
            password='testpass123',
            is_superuser=True,
            is_staff=True,
        )
        response = self.client.get(reverse('user_delete', args=[loadtest.id]))
        self.assertRedirects(response, reverse('user_list'))
        self.assertFalse(User.objects.filter(username='loadtest').exists())


class UserAddFormTests(TestCase):
    def setUp(self):
        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={'module_permissions': {MODULE_HRM: {'view': True, 'edit': True}}},
        )
        self.admin = User.objects.create_user(username='hr_add', password='testpass123', is_staff=True)
        Profile.objects.filter(user=self.admin).update(
            role=ROLE_DIRECTOR,
            full_name='HR Add',
            is_employed=True,
            permission_group=None,
        )
        self.client = Client(HTTP_HOST='testserver')
        self.client.force_login(self.admin)

    def test_user_add_get_prefills_password(self):
        response = self.client.get(reverse('user_add'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="password"')

    def test_user_add_get_prefills_probation_defaults(self):
        response = self.client.get(reverse('user_add'))
        self.assertEqual(response.status_code, 200)
        form = response.context['form']
        self.assertEqual(form.initial.get('job_position'), 'Nhân viên thử việc')
        self.assertEqual(form.initial.get('job_title'), 'Nhân viên thử việc')
        self.assertEqual(form.initial.get('role'), ROLE_EMPLOYEE)
        self.assertTrue(form.initial.get('join_date'))
        self.assertTrue(form.initial.get('employee_code'))

    def test_user_suggest_username_api(self):
        response = self.client.get(
            reverse('user_suggest_username'),
            {'full_name': 'Nguyễn Thành Nam'},
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['username'], 'nam.nt')
        self.assertEqual(data['email'], 'nam.nt@justplay.vn')
        self.assertTrue(data['employee_code'])


    def test_user_add_get_has_avatar_and_tabs(self):
        response = self.client.get(reverse('user_add'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="id_avatar"')
        self.assertContains(response, 'data-tab="permissions"')
        self.assertContains(response, 'jp-user-form-tab-panels')

    def test_user_add_post_success_without_email(self):
        response = self.client.post(reverse('user_add'), {
            'username': 'newstaff01',
            'password': 'TestPass1',
            'email': '',
            'full_name': 'Nguyễn Văn Mới',
            'role': ROLE_EMPLOYEE,
            'is_employed': '1',
            'concurrent-TOTAL_FORMS': '1',
            'concurrent-INITIAL_FORMS': '0',
            'concurrent-MIN_NUM_FORMS': '0',
            'concurrent-MAX_NUM_FORMS': '1000',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(User.objects.filter(username='newstaff01').exists())

    def test_user_add_post_shows_errors_when_missing_name(self):
        response = self.client.post(reverse('user_add'), {
            'username': 'x',
            'password': 'TestPass1',
            'full_name': '',
            'role': ROLE_EMPLOYEE,
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Không lưu được')
        self.assertContains(response, 'Họ và tên')


class UserSearchTests(TestCase):
    def setUp(self):
        dept = Department.objects.create(name='Phòng May', sort_order=1)
        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={'module_permissions': {MODULE_HRM: {'view': True, 'edit': True}}},
        )
        self.admin = User.objects.create_user(username='hr_admin', password='testpass123', is_staff=True)
        Profile.objects.filter(user=self.admin).update(
            role=ROLE_DIRECTOR,
            full_name='HR Admin',
            is_employed=True,
            permission_group=None,
        )
        self.target = User.objects.create_user(username='annt', password='testpass123', first_name='An')
        Profile.objects.filter(user=self.target).update(
            full_name='Nguyễn Văn An',
            employee_code='NV12345',
            department=dept,
            is_employed=True,
        )
        self.other = User.objects.create_user(username='other_user', password='testpass123')
        Profile.objects.filter(user=self.other).update(full_name='Trần Văn B', is_employed=True)
        self.client = Client(HTTP_HOST='testserver')
        self.client.force_login(self.admin)

    def test_search_by_employee_code_finds_user(self):
        response = self.client.get(reverse('user_list'), {'q': 'NV12345'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nguyễn Văn An')
        self.assertNotContains(response, 'Trần Văn B')

    def test_search_by_username_finds_user(self):
        response = self.client.get(reverse('user_list'), {'q': 'annt'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nguyễn Văn An')

    def test_user_list_hides_system_admin_account(self):
        User.objects.create_user(username='admin', password='testpass123', is_superuser=True)
        response = self.client.get(reverse('user_list'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, '>admin<')
        self.assertNotContains(response, '@admin')

    def test_user_list_filter_by_department(self):
        dept_may = Department.objects.get(name='Phòng May')
        response = self.client.get(reverse('user_list'), {'department': str(dept_may.pk)})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nguyễn Văn An')
        self.assertNotContains(response, 'Trần Văn B')

    def test_user_list_filter_unassigned_department(self):
        response = self.client.get(reverse('user_list'), {'department': 'none'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Trần Văn B')
        self.assertNotContains(response, 'Nguyễn Văn An')

    def test_user_list_filter_by_division_and_position(self):
        dept = Department.objects.get(name='Phòng May')
        div = Division.objects.create(name='Tổ QC', department=dept, sort_order=1)
        Profile.objects.filter(user=self.target).update(
            division=div,
            job_position='Kiểm tra chất lượng',
        )
        response = self.client.get(reverse('user_list'), {
            'division': str(div.pk),
            'position': 'Kiểm tra chất lượng',
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Nguyễn Văn An')
        self.assertNotContains(response, 'Trần Văn B')

        response2 = self.client.get(reverse('user_list'), {'division': str(div.pk)})
        self.assertContains(response2, 'Nguyễn Văn An')
        self.assertNotContains(response2, 'Trần Văn B')

    def test_user_list_filter_by_employment_status(self):
        Profile.objects.filter(user=self.other).update(is_employed=False)
        active = self.client.get(reverse('user_list'), {'status': 'active'})
        self.assertEqual(active.status_code, 200)
        self.assertContains(active, 'Nguyễn Văn An')
        self.assertNotContains(active, 'Trần Văn B')

        inactive = self.client.get(reverse('user_list'), {'status': 'inactive'})
        self.assertContains(inactive, 'Trần Văn B')
        self.assertNotContains(inactive, 'Nguyễn Văn An')

    def test_user_list_sort_by_code_uses_numeric_order(self):
        from hrm.user_search import apply_user_list_sort, exclude_hidden_hrm_users

        samples = (
            ('code_101', '101'),
            ('code_001', '001'),
            ('code_11', '11'),
            ('code_055', '055'),
        )
        for username, code in samples:
            user = User.objects.create_user(username=username, password='testpass123')
            Profile.objects.filter(user=user).update(
                employee_code=code,
                full_name=f'NV {code}',
                is_employed=True,
            )

        qs = apply_user_list_sort(
            exclude_hidden_hrm_users(User.objects.select_related('profile')),
            'code',
            'asc',
        )
        codes = [
            u.profile.employee_code
            for u in qs
            if u.profile.employee_code in {'001', '11', '055', '101'}
        ]
        self.assertEqual(codes[:4], ['001', '11', '055', '101'])

    def test_user_list_sortable_column_headers(self):
        response = self.client.get(reverse('user_list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="userStatusFilter"')
        self.assertNotContains(response, 'id="userSort"')
        self.assertContains(response, 'jp-table-sort-link')
        self.assertContains(response, 'sort=code&amp;dir=asc')
        self.assertContains(response, 'sort=name&amp;dir=desc')
        self.assertContains(response, 'is-active is-asc')
        self.assertContains(response, 'Nhóm quyền')
        self.assertNotContains(response, 'data-label="Ngày vào"')

    def test_user_edit_redirect_preserves_list_filters(self):
        dept = Department.objects.get(name='Phòng May')
        edit_url = reverse('user_edit', args=[self.target.id]) + f'?department={dept.pk}&status=active&q=NV'
        self.client.get(edit_url)
        response = self.client.post(edit_url, {
            'list_return_query': f'department={dept.pk}&status=active&q=NV',
            'username': 'annt',
            'email': 'annt@justplay.vn',
            'full_name': 'Nguyễn Văn An',
            'employee_code': 'NV12345',
            'role': ROLE_EMPLOYEE,
            'is_employed': '1',
            'concurrent-TOTAL_FORMS': '0',
            'concurrent-INITIAL_FORMS': '0',
            'concurrent-MIN_NUM_FORMS': '0',
            'concurrent-MAX_NUM_FORMS': '1000',
        })
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('user_list'), response.url)
        self.assertIn(f'department={dept.pk}', response.url)
        self.assertIn('status=active', response.url)
        self.assertIn('q=NV', response.url)

    def test_display_title_filter(self):
        from hrm.templatetags.hrm_extras import display_title

        self.assertEqual(display_title('TRẦN NHÂN ĐỨC'), 'Trần Nhân Đức')
        self.assertEqual(display_title('SẢN XUẤT'), 'Sản Xuất')
        self.assertEqual(display_title('admin'), 'admin')
        self.assertEqual(display_title('IT / CNTT'), 'It / Cntt')

    def test_user_list_displays_title_case_for_name_and_org(self):
        dept = Department.objects.create(name='PHÒNG TEST UPPER', sort_order=99)
        division = Division.objects.create(name='CẮT, TRẢI VẢI', department=dept, sort_order=1)
        Profile.objects.filter(user=self.target).update(
            full_name='PHẠM THỊ NGUYỄN',
            department=dept,
            division=division,
        )

        response = self.client.get(reverse('user_list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Phạm Thị Nguyễn')
        self.assertNotContains(response, 'PHẠM THỊ NGUYỄN')
        self.assertContains(response, 'Phòng Test Upper')
        self.assertContains(response, 'Cắt, Trải Vải')

    def test_hr_edit_can_update_other_user_avatar(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        image = SimpleUploadedFile(
            'staff.jpg',
            ProfileAvatarTests._sample_jpeg_bytes(),
            content_type='image/jpeg',
        )
        response = self.client.post(reverse('user_edit', args=[self.target.id]), {
            'username': 'annt',
            'email': 'annt@justplay.vn',
            'full_name': 'Nguyễn Văn An',
            'employee_code': 'NV12345',
            'role': ROLE_EMPLOYEE,
            'is_employed': '1',
            'avatar': image,
            'concurrent-TOTAL_FORMS': '0',
            'concurrent-INITIAL_FORMS': '0',
            'concurrent-MIN_NUM_FORMS': '0',
            'concurrent-MAX_NUM_FORMS': '1000',
        })
        self.assertEqual(response.status_code, 302)
        self.target.profile.refresh_from_db()
        self.assertTrue(self.target.profile.avatar)
        self.assertIn('staff', self.target.profile.avatar.name)


class MediaCleanupTests(TestCase):
    def test_normalize_media_relative_path(self):
        from hrm.media_cleanup import normalize_media_relative_path

        self.assertEqual(normalize_media_relative_path('avatars/2026/05/me.jpg'), 'avatars/2026/05/me.jpg')
        self.assertEqual(normalize_media_relative_path('/media/avatars/x.jpg'), 'avatars/x.jpg')
        self.assertEqual(
            normalize_media_relative_path('https://portal.justplay.vn/media/documents/pdf/a.pdf'),
            'documents/pdf/a.pdf',
        )

    def test_cleanup_orphan_media_dry_run(self):
        import tempfile
        from pathlib import Path

        from django.conf import settings
        from django.test import override_settings

        from hrm.media_cleanup import cleanup_orphan_media

        with tempfile.TemporaryDirectory() as tmp:
            media_root = Path(tmp)
            orphan = media_root / 'avatars' / '2026' / '05'
            orphan.mkdir(parents=True)
            orphan_file = orphan / 'old.jpg'
            orphan_file.write_bytes(b'orphan')

            with override_settings(MEDIA_ROOT=str(media_root)):
                result = cleanup_orphan_media(dry_run=True)
                self.assertEqual(result['orphan_count'], 1)
                self.assertTrue(orphan_file.exists())

                result = cleanup_orphan_media(dry_run=False)
                self.assertEqual(result['removed_count'], 1)
                self.assertFalse(orphan_file.exists())


class OrgStructureTreemapTests(TestCase):
    def setUp(self):
        from hrm.permissions import ROLE_DIRECTOR

        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={'module_permissions': {MODULE_HRM: {'view': True, 'edit': True}}},
        )
        self.admin = User.objects.create_user(
            username='org_admin',
            password='test',
            is_staff=True,
        )
        Profile.objects.filter(user=self.admin).update(
            role=ROLE_DIRECTOR,
            full_name='Org Admin',
            is_employed=True,
        )
        self.dept = Department.objects.create(name='ORG-DEPT-A', sort_order=1)
        self.dept_b = Department.objects.create(name='ORG-DEPT-B', sort_order=2)
        Division.objects.create(name='ORG-DIV-1', department=self.dept, sort_order=1)
        Division.objects.create(name='ORG-DIV-2', department=self.dept, sort_order=2)
        Division.objects.create(name='ORG-DIV-ORPHAN', department=None, sort_order=3)
        self.client.force_login(self.admin)

    def test_org_structure_treemap_renders_hierarchy(self):
        response = self.client.get(reverse('org_structure'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'jp-org-tree-data')
        self.assertContains(response, 'subtitle')
        self.assertContains(response, 'ORG-DEPT-A')
        self.assertContains(response, 'ORG-DIV-1')
        self.assertContains(response, 'ORG-DIV-ORPHAN')
        self.assertContains(response, 'jp-org-tree-mount')
        self.assertContains(response, 'jp-org-tree-data')
        self.assertContains(response, 'ORG-DEPT-A')
        self.assertContains(response, 'org_tree.js')
        self.assertContains(response, 'org-manage-panel')
        self.assertContains(response, 'org-tab-depts')
        self.assertContains(response, 'org-tab-divs')
        self.assertContains(response, 'Cập nhật sơ đồ')
        self.assertContains(response, 'Sơ đồ tổ chức')
        self.assertNotContains(response, 'Quản lý qua bảng')
        self.assertNotContains(response, 'Sơ đồ cây ngang')
        self.assertContains(response, 'jp-org-chart-headers-bar')
        self.assertContains(response, 'Phòng ban')
        self.assertContains(response, 'jp-org-urls-data')
        self.assertContains(response, 'positionAdd')

    def test_division_form_rejects_duplicate_in_same_department(self):
        from hrm.forms import DivisionForm

        Division.objects.create(name='QC-SHARED', department=self.dept)
        form = DivisionForm({
            'department': self.dept.pk,
            'name': 'QC-SHARED',
            'sort_order': 0,
            'is_active': True,
        })
        self.assertFalse(form.is_valid())
        form_b = DivisionForm({
            'department': self.dept_b.pk,
            'name': 'QC-SHARED',
            'sort_order': 0,
            'is_active': True,
        })
        self.assertTrue(form_b.is_valid())

    def test_org_tree_hides_system_admin_account(self):
        from hrm.models import DivisionPosition

        div = Division.objects.get(name='ORG-DIV-1')
        DivisionPosition.objects.create(
            division=div,
            department=self.dept,
            name='ORG-ADMIN-POS',
        )
        admin_user = User.objects.create_user(
            username='admin',
            password='x',
            first_name='System',
            is_superuser=True,
        )
        Profile.objects.update_or_create(
            user=admin_user,
            defaults={
                'full_name': 'System Admin',
                'employee_code': 'ADM01',
                'department': self.dept,
                'division': div,
                'job_position': 'ORG-ADMIN-POS',
                'is_employed': True,
            },
        )
        response = self.client.get(reverse('org_structure'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'System Admin')
        self.assertNotContains(response, 'ADM01')
        self.assertNotContains(response, '"level": "employee", "id": %d' % admin_user.pk)

    def test_division_delete_ignores_hidden_admin_profile(self):
        from hrm.models import DivisionPosition

        div = Division.objects.get(name='ORG-DIV-1')
        DivisionPosition.objects.create(
            division=div,
            department=self.dept,
            name='QC-DEL',
        )
        admin_user = User.objects.create_user(
            username='admin',
            password='x',
            is_superuser=True,
        )
        Profile.objects.update_or_create(
            user=admin_user,
            defaults={
                'full_name': 'Hidden Admin',
                'department': self.dept,
                'division': div,
                'job_position': 'QC-DEL',
                'is_employed': True,
            },
        )
        self.assertEqual(div.employee_count, 0)
        response = self.client.get(reverse('division_delete', args=[div.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Xóa bộ phận')
        self.assertNotContains(response, 'đang có <strong>1</strong> nhân viên')

    def test_org_tree_employee_has_avatar_url_field(self):
        from hrm.org_structure import build_org_tree, build_org_treemap

        div = Division.objects.get(name='ORG-DIV-1')
        user = User.objects.create_user(username='orgav1', password='x')
        Profile.objects.update_or_create(
            user=user,
            defaults={
                'full_name': 'NV Avatar',
                'department': self.dept,
                'division': div,
                'job_position': 'ORG-EMP-POS',
                'is_employed': True,
            },
        )
        tree = build_org_tree(build_org_treemap())
        dept = next(c for c in tree['children'] if c.get('id') == self.dept.pk)
        div_node = next(c for c in dept['children'] if c.get('name') == 'ORG-DIV-1')
        emp = next(
            (c for pos in div_node['children'] for c in pos.get('children', [])
             if c.get('name') == 'NV Avatar'),
            None,
        )
        self.assertIsNotNone(emp)
        self.assertIn('avatar_url', emp)

    def test_position_node_limits_chart_preview(self):
        from hrm.models import DivisionPosition
        from hrm.org_structure import ORG_CHART_EMPLOYEE_PREVIEW, build_org_tree, build_org_treemap

        div = Division.objects.get(name='ORG-DIV-1')
        DivisionPosition.objects.create(
            division=div,
            department=self.dept,
            name='ORG-PREV-POS',
        )
        for i in range(ORG_CHART_EMPLOYEE_PREVIEW + 3):
            user = User.objects.create_user(username=f'prev{i}', password='x')
            Profile.objects.update_or_create(
                user=user,
                defaults={
                    'full_name': f'Preview {i}',
                    'department': self.dept,
                    'division': div,
                    'job_position': 'ORG-PREV-POS',
                    'is_employed': True,
                },
            )
        tree = build_org_tree(build_org_treemap())
        dept = next(c for c in tree['children'] if c.get('id') == self.dept.pk)
        div_node = next(c for c in dept['children'] if c.get('name') == 'ORG-DIV-1')
        pos = next((c for c in div_node['children'] if c.get('name') == 'ORG-PREV-POS'), None)
        self.assertIsNotNone(pos)
        self.assertEqual(pos['employee_total'], ORG_CHART_EMPLOYEE_PREVIEW + 3)
        self.assertTrue(pos['has_more_employees'])
        self.assertEqual(len(pos['children']), ORG_CHART_EMPLOYEE_PREVIEW)
        self.assertEqual(len(pos['employees_all']), ORG_CHART_EMPLOYEE_PREVIEW + 3)

    def test_org_tree_includes_employee_children(self):
        from hrm.models import DivisionPosition

        div = Division.objects.get(name='ORG-DIV-1')
        DivisionPosition.objects.create(
            division=div,
            department=self.dept,
            name='ORG-EMP-POS',
        )
        user = User.objects.create_user(username='orgemp1', password='x', first_name='Emp One')
        Profile.objects.update_or_create(
            user=user,
            defaults={
                'full_name': 'Nhân viên Org Test',
                'employee_code': 'OE1',
                'department': self.dept,
                'division': div,
                'job_position': 'ORG-EMP-POS',
                'is_employed': True,
            },
        )
        response = self.client.get(reverse('org_structure'))
        self.assertContains(response, '"level": "employee"')
        self.assertContains(response, 'OE1')
        self.assertContains(response, 'ORG-EMP-POS')

    def test_division_position_in_org_tree(self):
        from hrm.models import DivisionPosition

        div = Division.objects.get(name='ORG-DIV-1')
        DivisionPosition.objects.create(
            division=div,
            department=self.dept,
            name='ORG-POS-QC',
        )
        response = self.client.get(reverse('org_structure'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ORG-POS-QC')

    def test_import_template_multisheet(self):
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(reverse('user_download_template'))
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        self.assertIn('Nhap_lieu', wb.sheetnames)
        self.assertIn('Phong_ban', wb.sheetnames)
        self.assertIn('Huong_dan', wb.sheetnames)

    def test_org_position_add_get(self):
        div = Division.objects.get(name='ORG-DIV-1')
        url = reverse('org_position_add') + f'?division={div.pk}&department={self.dept.pk}'
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Thêm vị trí')

    def test_divisions_for_department_matches_org_and_profiles(self):
        from hrm.org_structure import divisions_for_department

        div1 = Division.objects.get(name='ORG-DIV-1')
        div_orphan = Division.objects.get(name='ORG-DIV-ORPHAN')
        div_inactive = Division.objects.create(
            name='ORG-DIV-INACTIVE',
            department=self.dept,
            is_active=False,
        )
        div_wrong_fk = Division.objects.create(
            name='ORG-DIV-WRONG-FK',
            department=self.dept_b,
            is_active=True,
        )
        emp = User.objects.create_user(username='div_filter_emp', password='x')
        Profile.objects.update_or_create(
            user=emp,
            defaults={
                'full_name': 'NV Wrong FK',
                'department': self.dept,
                'division': div_wrong_fk,
                'is_employed': True,
            },
        )

        ids = set(divisions_for_department(self.dept.pk).values_list('pk', flat=True))
        self.assertIn(div1.pk, ids)
        self.assertIn(div_orphan.pk, ids)
        self.assertIn(div_inactive.pk, ids)
        self.assertIn(div_wrong_fk.pk, ids)

    def test_user_add_form_includes_divisions_allowed_json(self):
        url = reverse('user_add')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'jpDivisionsByDept')
        self.assertContains(response, str(self.dept.pk))

    def test_user_list_filter_divisions_strict_by_department(self):
        from hrm.org_structure import divisions_for_user_list_filter

        div_a = Division.objects.create(name='UL-DIV-A', department=self.dept, is_active=True)
        div_b = Division.objects.create(name='UL-DIV-B', department=self.dept_b, is_active=True)
        ids = set(divisions_for_user_list_filter(self.dept.pk).values_list('pk', flat=True))
        self.assertIn(div_a.pk, ids)
        self.assertNotIn(div_b.pk, ids)

    def test_user_list_page_filter_assets(self):
        self.client.force_login(self.admin)
        response = self.client.get(reverse('user_list'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'jpUserListDivisionsMap')
        self.assertContains(response, 'jpUserListPositionsCascade')
        self.assertNotContains(response, 'Lọc tự động')
        self.assertNotContains(response, 'Dữ liệu nhân sự')

        filtered = self.client.get(reverse('user_list') + f'?department={self.dept.pk}')
        self.assertEqual(filtered.status_code, 200)
        html = filtered.content.decode()
        div_select_start = html.index('id="userDivisionFilter"')
        div_select_end = html.index('</select>', div_select_start)
        division_select = html[div_select_start:div_select_end]
        self.assertIn('ORG-DIV-1', division_select)
        self.assertIn('ORG-DIV-2', division_select)
        self.assertNotIn('ORG-DIV-ORPHAN', division_select)

    def test_org_position_edit_moves_employees_to_other_division(self):
        from hrm.models import DivisionPosition
        from hrm.org_structure import build_org_tree, build_org_treemap

        div1 = Division.objects.get(name='ORG-DIV-1')
        div2 = Division.objects.get(name='ORG-DIV-2')
        pos = DivisionPosition.objects.create(
            division=div1,
            department=self.dept,
            name='May chuyen BP',
        )
        emp = User.objects.create_user(username='move_pos_emp', password='x')
        profile, _ = Profile.objects.update_or_create(
            user=emp,
            defaults={
                'full_name': 'NV Chuyen BP',
                'employee_code': 'MVBP1',
                'department': self.dept,
                'division': div1,
                'job_position': 'May chuyen BP',
                'is_employed': True,
            },
        )
        url = reverse('org_position_edit', kwargs={'pk': pos.pk})
        response = self.client.post(url, {
            'division': div2.pk,
            'name': 'May chuyen BP',
            'sort_order': pos.sort_order,
            'is_active': True,
        })
        self.assertEqual(response.status_code, 302)
        profile.refresh_from_db()
        self.assertEqual(profile.division_id, div2.pk)
        self.assertEqual(profile.department_id, self.dept.pk)
        self.assertEqual(profile.job_position, 'May chuyen BP')

        tree = build_org_tree(build_org_treemap())
        div2_node = None
        for dept in tree.get('children', []):
            for div in dept.get('children', []):
                if div.get('id') == div2.pk:
                    div2_node = div
                    break
        self.assertIsNotNone(div2_node)
        pos_names = [c['name'] for c in div2_node.get('children', []) if c.get('level') == 'position']
        self.assertIn('May chuyen BP', pos_names)
        emp_nodes = []
        for p in div2_node.get('children', []):
            if p.get('level') == 'position' and p.get('name') == 'May chuyen BP':
                emp_nodes = p.get('children', [])
                break
        self.assertTrue(any(n.get('employee_code') == 'MVBP1' for n in emp_nodes))

    def test_org_tree_department_head_as_subtitle(self):
        from hrm.org_structure import (
            ORG_DEPARTMENT_HEAD_LABEL,
            ORG_DEPARTMENT_HEAD_PREFIX,
            build_org_tree,
            build_org_treemap,
        )

        treemap = build_org_treemap()
        tree = build_org_tree(treemap)
        dept_node = next(
            c for c in tree['children']
            if c.get('level') == 'department' and c.get('id') == self.dept.pk
        )
        self.assertFalse(dept_node.get('has_head'))
        self.assertNotIn('subtitle', dept_node)
        self.assertEqual(dept_node['children'][0]['level'], 'division')

        user = User.objects.create_user(username='depthead1', password='x', first_name='Head')
        Profile.objects.update_or_create(
            user=user,
            defaults={
                'full_name': 'Nguyễn Thành An',
                'employee_code': 'TH1',
                'department': self.dept,
                'division': None,
                'job_position': ORG_DEPARTMENT_HEAD_LABEL,
                'is_employed': True,
            },
        )
        tree2 = build_org_tree(build_org_treemap())
        dept2 = next(
            c for c in tree2['children']
            if c.get('level') == 'department' and c.get('id') == self.dept.pk
        )
        self.assertTrue(dept2.get('has_head'))
        self.assertIn('Nguyễn Thành An', dept2['subtitle'])
        self.assertTrue(dept2['subtitle'].startswith(ORG_DEPARTMENT_HEAD_PREFIX))

        response = self.client.get(reverse('org_structure'))
        self.assertContains(response, f'"head_user_id": {user.pk}')
        self.assertContains(response, 'has_head')

    def test_org_tree_division_head_subtitle_and_urls(self):
        from hrm.org_structure import (
            ORG_DIVISION_HEAD_LABEL,
            ORG_DIVISION_HEAD_PREFIX,
            build_org_tree,
            build_org_treemap,
        )
        from hrm.permissions import ROLE_DIVISION_HEAD

        div = Division.objects.get(name='ORG-DIV-1')
        user = User.objects.create_user(username='divhead1', password='x')
        Profile.objects.update_or_create(
            user=user,
            defaults={
                'full_name': 'Trưởng BP Test',
                'department': self.dept,
                'division': div,
                'role': ROLE_DIVISION_HEAD,
                'job_position': ORG_DIVISION_HEAD_LABEL,
                'is_employed': True,
            },
        )
        tree = build_org_tree(build_org_treemap())
        dept_node = next(
            c for c in tree['children']
            if c.get('level') == 'department' and c.get('id') == self.dept.pk
        )
        div_node = next(c for c in dept_node['children'] if c.get('id') == div.pk)
        self.assertTrue(div_node.get('has_head'))
        self.assertIn('Trưởng BP Test', div_node['subtitle'])
        self.assertTrue(div_node['subtitle'].startswith(ORG_DIVISION_HEAD_PREFIX))

        response = self.client.get(reverse('org_structure'))
        self.assertContains(response, 'divHeadAssign')
        self.assertContains(response, 'directorAssign')
        self.assertContains(response, ROLE_DIVISION_HEAD)

    def test_resolve_division_scoped_to_department(self):
        from hrm.choices import resolve_division

        Division.objects.create(name='QC-SCOPE', department=self.dept)
        div_b = Division.objects.create(name='QC-SCOPE', department=self.dept_b)
        div = resolve_division('QC-SCOPE', department=self.dept_b)
        self.assertEqual(div.pk, div_b.pk)

    def test_auto_sort_order_on_create(self):
        from hrm.forms import DepartmentForm, DivisionForm, DivisionPositionForm
        from hrm.models import DivisionPosition

        auto_dept = DepartmentForm().fields['sort_order'].initial
        self.assertIsNotNone(auto_dept)
        form = DepartmentForm({'name': 'AUTO-D3', 'sort_order': auto_dept, 'is_active': True})
        self.assertTrue(form.is_valid(), form.errors)
        d3 = form.save()
        self.assertEqual(d3.sort_order, auto_dept)

        div = Division.objects.get(name='ORG-DIV-1')
        Division.objects.filter(department=self.dept).exclude(pk=div.pk).delete()
        Division.objects.create(name='AUTO-DIV-A', department=self.dept, sort_order=0)
        Division.objects.create(name='AUTO-DIV-B', department=self.dept, sort_order=2)
        auto_div = DivisionForm(initial={'department': self.dept.pk}).fields['sort_order'].initial
        self.assertIsNotNone(auto_div)
        dform = DivisionForm({
            'department': self.dept.pk,
            'name': 'AUTO-DIV-C',
            'sort_order': auto_div,
            'is_active': True,
        }, initial={'department': self.dept.pk})
        self.assertTrue(dform.is_valid(), dform.errors)
        new_div = dform.save()
        self.assertEqual(new_div.sort_order, auto_div)

        DivisionPosition.objects.filter(division=div).delete()
        DivisionPosition.objects.create(
            division=div, department=self.dept, name='AUTO-P1', sort_order=0,
        )
        auto_pos = DivisionPositionForm(initial={'division': div.pk}).fields['sort_order'].initial
        self.assertIsNotNone(auto_pos)
        pform = DivisionPositionForm({
            'division': div.pk,
            'name': 'AUTO-P2',
            'sort_order': auto_pos,
            'is_active': True,
        }, initial={'division': div.pk})
        self.assertTrue(pform.is_valid(), pform.errors)
        pos = pform.save()
        self.assertEqual(pos.sort_order, auto_pos)

        # Sửa thủ công thứ tự thì giữ nguyên
        dform_manual = DivisionForm({
            'department': self.dept.pk,
            'name': 'AUTO-DIV-MANUAL',
            'sort_order': 99,
            'is_active': True,
        }, initial={'department': self.dept.pk})
        self.assertTrue(dform_manual.is_valid())
        manual = dform_manual.save()
        self.assertEqual(manual.sort_order, 99)


class ExcelHrFieldsTests(TestCase):
    def test_export_row_includes_role_permission_and_status(self):
        from hrm.choices import EXCEL_ALL_HEADERS, user_to_excel_row

        dept = Department.objects.create(name='EXCEL-DEPT', sort_order=1)
        group = PermissionGroup.objects.create(name='EXCEL-GROUP', slug='excel-group')
        user = User.objects.create_user(username='excel_exp', password='x', first_name='Excel User')
        Profile.objects.filter(user=user).update(
            full_name='Excel User',
            department=dept,
            role=ROLE_TEAM_LEADER,
            permission_group=group,
            is_employed=False,
        )
        user = User.objects.select_related('profile', 'profile__permission_group').get(pk=user.pk)
        row = user_to_excel_row(user)
        for header in EXCEL_ALL_HEADERS:
            self.assertIn(header, row)
        self.assertEqual(row['Vai trò HT'], 'Tổ trưởng')
        self.assertEqual(row['Nhóm quyền'], 'EXCEL-GROUP')
        self.assertEqual(row['Trạng thái'], 'Nghỉ việc')

    def test_import_defaults_include_role_permission_and_status(self):
        from hrm.choices import profile_defaults_from_import

        dept = Department.objects.create(name='EXCEL-IMP-DEPT', sort_order=1)
        group = PermissionGroup.objects.create(name='EXCEL-IMP-GROUP', slug='excel-imp-group')
        defaults = profile_defaults_from_import({
            'full_name': 'Import Test',
            'department': dept.name,
            'role': 'Giám đốc',
            'permission_group': 'EXCEL-IMP-GROUP',
            'is_employed': 'Nghỉ việc',
        })
        self.assertEqual(defaults['role'], ROLE_DIRECTOR)
        self.assertEqual(defaults['permission_group'], group)
        self.assertFalse(defaults['is_employed'])

    def test_import_template_has_new_columns(self):
        from io import BytesIO

        from openpyxl import load_workbook

        RoleModulePermission.objects.update_or_create(
            role=ROLE_DIRECTOR,
            defaults={'module_permissions': {MODULE_HRM: {'view': True, 'edit': True}}},
        )
        admin = User.objects.create_user(username='excel_tpl', password='x', is_staff=True)
        Profile.objects.filter(user=admin).update(role=ROLE_DIRECTOR, is_employed=True)
        client = Client(HTTP_HOST='testserver')
        client.force_login(admin)

        response = client.get(reverse('user_download_template'))
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in wb['Nhap_lieu'][1]]
        self.assertIn('Vai trò HT', headers)
        self.assertIn('Nhóm quyền', headers)
        self.assertIn('Trạng thái', headers)

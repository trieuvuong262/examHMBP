from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from django.contrib.auth.models import User
from hrm.permissions import (
    ROLE_DIRECTOR,
    ROLE_EMPLOYEE,
    ROLE_TEAM_LEADER,
    SUBORDINATE_MANAGER_ROLES,
    can_manage_kpi_for_others,
    get_profile,
    is_gm,
    user_role,
)
from .models import KpiPeriod, YearlyKpi, YearlyKpiItem
import openpyxl
from django.http import HttpResponse
import datetime
from PortalJustPlay.list_search import apply_combined_search, get_search_query
from PortalJustPlay.pagination import paginate_queryset


def _period_title(period_type: str) -> str:
    return dict(KpiPeriod.PERIOD_CHOICES).get(period_type, period_type)


def _get_or_create_period(year: int, period_type: str) -> KpiPeriod:
    title = _period_title(period_type)
    period = KpiPeriod.objects.filter(year=year, period_type=period_type).order_by('id').first()
    if period:
        if not period.title:
            period.title = title
            period.save(update_fields=['title'])
        return period
    return KpiPeriod.objects.create(
        year=year,
        period_type=period_type,
        title=title,
    )
# ========================================================
# 1. TRANG DANH SÁCH KPI (DASHBOARD CHÍNH)
# ========================================================
@login_required
def kpi_list_view(request):
    search_query = get_search_query(request)
    my_kpis_qs = YearlyKpi.objects.filter(employee=request.user).order_by('-year')
    team_kpis_qs = YearlyKpi.objects.filter(direct_manager=request.user).select_related(
        'employee__profile',
    ).order_by('-year')
    
    role = user_role(request.user)

    # Nếu là GM hoặc Admin thì thấy toàn bộ
    if role == ROLE_DIRECTOR or request.user.is_superuser:
        team_kpis_qs = YearlyKpi.objects.exclude(employee=request.user).select_related(
            'employee__profile',
        ).order_by('-year')

    def _kpi_search(qs):
        if not search_query:
            return qs
        return apply_combined_search(qs, search_query, lambda term: (
            Q(employee__username__icontains=term)
            | Q(employee__first_name__icontains=term)
            | Q(employee__last_name__icontains=term)
            | Q(employee__email__icontains=term)
            | Q(employee__profile__full_name__icontains=term)
            | Q(employee__profile__employee_code__icontains=term)
            | Q(direct_manager__username__icontains=term)
            | Q(direct_manager__profile__full_name__icontains=term)
            | (Q(year=int(term)) if term.isdigit() else Q())
        ))

    my_kpis_qs = _kpi_search(my_kpis_qs)
    team_kpis_qs = _kpi_search(team_kpis_qs)

    my_page, my_query_string = paginate_queryset(request, my_kpis_qs, page_param='my_page')
    team_page, team_query_string = paginate_queryset(request, team_kpis_qs, page_param='team_page')

    is_admin = request.user.is_superuser
    is_manager_or_gm = can_manage_kpi_for_others(request.user)

    # 2. XỬ LÝ QUYỀN ADMIN: Đóng/Mở kỳ đánh giá
    current_year = datetime.datetime.now().year
    admin_periods = []

    if is_admin:
        if request.method == 'POST' and 'toggle_period' in request.POST:
            p_type = request.POST.get('period_type')
            is_active = request.POST.get('is_active') == 'on' 
            
            period = _get_or_create_period(current_year, p_type)
            period.is_active = is_active
            period.save()
            messages.success(request, f"Đã {'MỞ' if is_active else 'ĐÓNG'} kỳ đánh giá {p_type}!")
            return redirect('kpi_list')
        
        standard_types = ['Q1', 'Q2', 'Q3', 'Q4', 'H1', 'H2', 'Y']
        for pt in standard_types:
            admin_periods.append(_get_or_create_period(current_year, pt))

    # TRUYỀN BIẾN is_manager_or_gm RA HTML
    return render(request, 'kpi/kpi_list.html', {
        'my_kpis': my_page.object_list,
        'my_page': my_page,
        'my_query_string': my_query_string,
        'team_kpis': team_page.object_list,
        'team_page': team_page,
        'team_query_string': team_query_string,
        'search_query': search_query,
        'is_admin': is_admin,
        'is_manager_or_gm': is_manager_or_gm,  # <--- Bắt buộc phải có dòng này nha ní!
        'admin_periods': admin_periods,
    })
    
@login_required
def kpi_detail_view(request, kpi_id):
    # 1. Lấy bảng KPI và các chỉ tiêu liên quan
    kpi_board = get_object_or_404(YearlyKpi, id=kpi_id)
    items = kpi_board.items.all().order_by('pillar')
    
    # 2. Lấy danh sách các kỳ ĐANG MỞ (Q1, Q2, Q3, Q4, H1, H2, Y)
    open_periods = KpiPeriod.objects.filter(year=kpi_board.year, is_active=True).values_list('period_type', flat=True)
    
    # 3. Phân quyền người dùng
    viewer_profile = get_profile(request.user)
    is_owner = (request.user == kpi_board.employee)
    is_manager = request.user == kpi_board.direct_manager
    if viewer_profile and viewer_profile.role in SUBORDINATE_MANAGER_ROLES:
        is_manager = is_manager or viewer_profile.subordinates.filter(
            pk=kpi_board.employee_id,
        ).exists()
    is_gm_user = is_gm(request.user) or request.user == kpi_board.general_manager
    
    # Nếu là Manager thì không cần làm Owner (để tránh xung đột logic nút bấm)
    if is_manager or is_gm_user:
        is_owner = (request.user == kpi_board.employee) # Giữ nguyên để biết lính tự xem bài mình

    if request.method == 'POST':
        target_period = request.POST.get('target_period') # Nhận Q1, Q2, H1, Y...
        action = request.POST.get('action', 'save') 

        if not target_period:
            messages.error(request, "Vui lòng chọn kỳ đánh giá cụ thể!")
            return redirect('kpi_detail', kpi_id=kpi_id)

        # Kiểm tra trạng thái đóng/mở kỳ từ Admin
        if target_period not in open_periods and action != 'save':
            messages.error(request, f"Kỳ đánh giá {target_period} hiện đang đóng, không thể nộp bài!")
            return redirect('kpi_detail', kpi_id=kpi_id)

        # Xác định tên trường trạng thái (VD: q1_status, y_status)
        status_attr = f"{target_period.lower()}_status"
        
        # Kiểm tra xem trường status có tồn tại trong Model không (đề phòng lỗi y_status chưa migrate)
        if not hasattr(kpi_board, status_attr):
            messages.error(request, f"Hệ thống chưa cấu hình trạng thái cho kỳ {target_period}!")
            return redirect('kpi_detail', kpi_id=kpi_id)
            
        current_status = getattr(kpi_board, status_attr)

        # 4. VÒNG LẶP LƯU ĐIỂM ĐỘNG
        for item in items:
            prefix = f"item_{item.id}_{target_period}_"
            
            try:
                # NHÂN VIÊN TỰ CHẤM
                if is_owner and current_status == 'self_evaluating':
                    val_self = request.POST.get(f"{prefix}self")
                    if val_self is not None:
                        setattr(item, f"{target_period.lower()}_self", float(val_self) if val_self.strip() != "" else None)
                
                # HOD CHẤM
                if is_manager and current_status == 'manager_evaluating':
                    val_mgr = request.POST.get(f"{prefix}mgr")
                    if val_mgr is not None:
                        setattr(item, f"{target_period.lower()}_mgr", float(val_mgr) if val_mgr.strip() != "" else None)
                    
                # GM CHỐT ĐIỂM
                if is_gm_user and current_status == 'general_evaluating':
                    val_gm = request.POST.get(f"{prefix}gm")
                    if val_gm is not None:
                        setattr(item, f"{target_period.lower()}_gm", float(val_gm) if val_gm.strip() != "" else None)
                
                item.save()
            except ValueError:
                messages.warning(request, f"Định dạng điểm không hợp lệ tại mục tiêu: {item.personal_objective}")
                continue

        # 5. CẬP NHẬT TRẠNG THÁI LUỒNG (Chỉ cập nhật khi bấm nút Nộp/Chốt)
        if action == 'submit_manager' and is_owner:
            setattr(kpi_board, status_attr, 'manager_evaluating')
            messages.success(request, f"Đã gửi đánh giá {target_period} cho Quản lý!")
            
        elif action == 'submit_gm' and is_manager:
            setattr(kpi_board, status_attr, 'general_evaluating')
            messages.success(request, f"Đã gửi kỳ {target_period} cho Quản lý cấp cao (GM)!")
            
        elif action == 'finish' and is_gm_user:
            setattr(kpi_board, status_attr, 'completed')
            messages.success(request, f"Chúc mừng! Đã chốt kết quả cuối cùng cho kỳ {target_period}!")
            
        else:
            messages.info(request, f"Đã lưu bản nháp dữ liệu kỳ {target_period}.")
            
        kpi_board.save()
        return redirect('kpi_detail', kpi_id=kpi_id)

    # 6. Render ra giao diện
    return render(request, 'kpi/kpi_form.html', {
        'kpi': kpi_board, 
        'items': items, 
        'open_periods': list(open_periods),
        'is_owner': is_owner, 
        'is_manager': is_manager, 
        'is_gm': is_gm_user
    })

# ========================================================
# 3. GIAO MỤC TIÊU NĂM (YEARLY SETUP)
# ========================================================
@login_required
def yearly_kpi_create(request):
    profile = get_profile(request.user)
    if not profile:
        messages.error(request, "Tài khoản chưa có hồ sơ nhân sự. Vui lòng liên hệ HR/IT.")
        return redirect('kpi_list')
    
    # 1. CHẶN ĐỨNG NHÂN VIÊN (Quy trình Top-Down)
    if profile.role == ROLE_EMPLOYEE and not request.user.is_superuser:
        messages.error(request, "Quyền truy cập bị từ chối! Chỉ Quản lý mới được thiết lập KPI năm.")
        return redirect('kpi_list')

    if profile.role == ROLE_TEAM_LEADER:
        target_employees = profile.subordinates.all()
    else:
        target_employees = User.objects.filter(is_active=True).exclude(id=request.user.id)

    hod_list = User.objects.filter(profile__role__in=SUBORDINATE_MANAGER_ROLES, is_active=True)
    gm_list = User.objects.filter(Q(profile__role=ROLE_DIRECTOR) | Q(is_superuser=True), is_active=True)
    
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        year = request.POST.get('year', timezone.now().year)
        eval_type = request.POST.get('eval_type', 'QUARTER')

        direct_manager_id = request.user.id if profile.role == ROLE_TEAM_LEADER else request.POST.get('direct_manager_id')
        general_manager_id = request.POST.get('general_manager_id')

        # Dùng update_or_create để nếu HOD lỡ tạo trùng năm thì ghi đè luôn, không bị sập web
        yearly_kpi, created = YearlyKpi.objects.update_or_create(
            employee_id=employee_id, 
            year=year,
            defaults={
                'direct_manager_id': direct_manager_id or None,
                'general_manager_id': general_manager_id or None,
                'eval_type': eval_type
            }
        )
        
        # Xóa các mục tiêu cũ của năm đó đi để add cái mới vào
        if not created:
            yearly_kpi.items.all().delete()

        # Bắt mảng dữ liệu từ Table trong HTML
        pillars = request.POST.getlist('pillar[]')
        objectives = request.POST.getlist('personal_objective[]')
        indicators = request.POST.getlist('kpi_indicator[]')
        weightages = request.POST.getlist('weightage[]')
        targets = request.POST.getlist('yearly_target[]')
        units = request.POST.getlist('unit[]')
        trends = request.POST.getlist('trend[]')

        for i in range(len(pillars)):
            if objectives[i].strip() != '':
                YearlyKpiItem.objects.create(
                    yearly_kpi=yearly_kpi, 
                    pillar=pillars[i], 
                    personal_objective=objectives[i],
                    kpi_indicator=indicators[i], 
                    weightage=float(weightages[i]) if weightages[i] else 0.0,
                    yearly_target=float(targets[i]) if targets[i] else 0.0, 
                    unit=units[i], 
                    trend=trends[i]
                )
                
        messages.success(request, f"Đã thiết lập Mục tiêu KPI Năm {year} thành công!")
        return redirect('kpi_list')

    return render(request, 'kpi/yearly_kpi_form.html', {
        'target_employees': target_employees,
        'hod_list': hod_list, 
        'gm_list': gm_list, 
        'is_hod': profile.role == ROLE_TEAM_LEADER,
    })


@login_required
def kpi_import_excel(request):
    profile = get_profile(request.user)
    if not profile:
        messages.error(request, "Tài khoản chưa có hồ sơ nhân sự. Vui lòng liên hệ HR/IT.")
        return redirect('kpi_list')
    if profile.role == ROLE_EMPLOYEE and not request.user.is_superuser:
        messages.error(request, "Bạn không có quyền này!")
        return redirect('kpi_list')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        direct_manager_id = request.POST.get('direct_manager_id')
        general_manager_id = request.POST.get('general_manager_id')
        eval_type = request.POST.get('eval_type', 'QUARTER')

        if profile.role == ROLE_TEAM_LEADER:
            direct_manager_id = request.user.id

        if not excel_file or not excel_file.name.endswith('.xlsx'):
            messages.error(request, "File không hợp lệ!")
            return redirect('kpi_import_excel')

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            processed_users_for_year = set() 

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row): continue 
                email, year, pillar, objective, indicator, weight, target, unit, trend = row[:9]
                
                employee = User.objects.filter(email=email).first()
                if not employee: continue
                    
                board, created = YearlyKpi.objects.update_or_create(
                    employee=employee, 
                    year=year,
                    defaults={
                        'direct_manager_id': direct_manager_id or None,
                        'general_manager_id': general_manager_id or None,
                        'eval_type': eval_type
                    }
                )
                
                if f"{employee.id}_{year}" not in processed_users_for_year:
                    board.items.all().delete()
                    processed_users_for_year.add(f"{employee.id}_{year}")

                YearlyKpiItem.objects.create(
                    yearly_kpi=board, pillar=pillar, personal_objective=objective,
                    kpi_indicator=indicator, weightage=float(weight or 0),
                    yearly_target=float(target or 0), unit=unit, trend=trend or 'HIGHER'
                )
            messages.success(request, "Import thành công!")
            return redirect('kpi_list')
        except Exception as e:
            messages.error(request, f"Lỗi: {str(e)}")
            return redirect('kpi_import_excel')

    # LẤY DANH SÁCH SẾP ĐỂ TRUYỀN RA GIAO DIỆN IMPORT
    hod_list = User.objects.filter(profile__role__in=SUBORDINATE_MANAGER_ROLES, is_active=True)
    gm_list = User.objects.filter(Q(profile__role=ROLE_DIRECTOR) | Q(is_superuser=True), is_active=True)

    return render(request, 'kpi/kpi_import.html', {
        'hod_list': hod_list,
        'gm_list': gm_list
    })
@login_required
def download_kpi_sample_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Email NV", "Năm", "Lĩnh vực", "Mục tiêu", "Chỉ số đo", "Trọng số (%)", "Chỉ tiêu", "Đơn vị", "Xu hướng"])
    ws.append(["luu.dao@justplay.vn", 2026, "FINANCE", "Mục tiêu mẫu", "Chỉ số mẫu", 20, 100, "%", "HIGHER"])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Mau_Import_KPI.xlsx'
    wb.save(response)
    return response
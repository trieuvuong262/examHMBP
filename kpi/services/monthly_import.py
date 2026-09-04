"""Parse / xuất Excel KPI tháng — chỉ bảng tiêu chí (không tiêu đề / tên / tháng)."""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


class KpiImportError(Exception):
    pass


@dataclass
class ParsedKpiRow:
    sort_order: int
    work_group: str
    weightage: float
    indicator: str
    level_fail: str
    level_pass: str
    level_exceed: str


@dataclass
class ParsedKpiSheet:
    rows: list[ParsedKpiRow]
    sheet_month: int | None = None
    sheet_year: int | None = None
    employee_name_hint: str = ''


_THIN = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
)

_HEADERS = [
    'STT',
    'Nhóm công việc',
    'Trọng số',
    'Tiêu chí đo lường (KPI)',
    'Mức Chưa đạt (<100%)',
    'Mức Đạt (100%)',
    'Mức Vượt (>100%)',
]

_SCORING_GUIDE_TITLE = 'HƯỚNG DẪN ĐÁNH GIÁ VÀ TÍNH ĐIỂM KPI'

_SCORING_GUIDE_BODY = """Nhân viên tự đánh giá vào cột Đánh giá thực tế (NV) và Điểm NV.
Quản lý đánh giá vào cột Đánh giá thực tế (QL) và Điểm QL.
Tổng điểm trên Portal ưu tiên điểm Quản lý; chưa có thì dùng điểm Nhân viên.

1. Thang điểm đánh giá
Mỗi tiêu chí sẽ được chấm trên thang điểm 10, tương ứng với mức độ hoàn thành công việc:

Từ 0 - 9 điểm (Mức Chưa đạt): Hoàn thành dưới 100% yêu cầu. Chấm điểm linh hoạt dựa trên tỷ lệ thực tế (ví dụ: hoàn thành 80% công việc thì chấm 8 điểm).

10 điểm (Mức Đạt): Hoàn thành đúng 100% yêu cầu công việc, đúng deadline và đạt chất lượng đề ra.

> 10 điểm (Mức Vượt - Điểm thưởng): nếu nhân sự hoàn thành vượt mức xuất sắc, mang lại giá trị lớn (tiết kiệm chi phí, vượt tiến độ), có thể chấm 11 hoặc 12 điểm cho tiêu chí đó.

2. Cách tính điểm
- Điểm thành phần = (Điểm đánh giá / 10) x Trọng số
- Tổng điểm KPI = Tổng các Điểm thành phần

3. Tiêu chí đánh giá tổng điểm
 - Nếu tổng điểm từ 0 - 89 điểm: Không đạt KPI
 - Nếu tổng điểm từ 90 - 100 điểm: Đạt KPI
 - Nếu tổng điểm từ 101 trở lên: Vượt KPI

4. Đánh giá thực tế
- Là nơi ghi nhận kết quả thực tế bằng các con số, sự việc hoặc bằng chứng cụ thể mà nhân sự đã đạt được trong tháng, nhằm đối chiếu trực tiếp với các «Tiêu chí đo lường (KPI)» đã đặt ra ban đầu.
"""


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _to_float_required(value: Any, *, row_no: int) -> float:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise KpiImportError(f'Dòng {row_no}: Trọng số không được để trống.')
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace('%', '').replace(',', '.')
    try:
        return float(text)
    except ValueError as exc:
        raise KpiImportError(f'Dòng {row_no}: Trọng số không hợp lệ ({value!r}).') from exc


def _find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 30) + 1):
        joined = ' '.join(_cell_str(ws.cell(r, c).value).lower() for c in range(1, 8))
        if 'tiêu chí' in joined and ('trọng số' in joined or 'trong so' in joined):
            return r
        if 'stt' in joined and 'tiêu chí' in joined:
            return r
    raise KpiImportError(
        'Không tìm thấy dòng tiêu đề. Cần các cột: STT, Nhóm công việc, Trọng số, '
        'Tiêu chí đo lường, Mức Chưa đạt, Mức Đạt, Mức Vượt.'
    )


def _merged_top_left_map(ws) -> dict[tuple[int, int], Any]:
    """Ô trong vùng merge → giá trị ô góc trên-trái."""
    mapping: dict[tuple[int, int], Any] = {}
    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        top_val = ws.cell(min_row, min_col).value
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                mapping[(r, c)] = top_val
    return mapping


def _cell_value(ws, row: int, col: int, merge_map: dict[tuple[int, int], Any]):
    if (row, col) in merge_map:
        return merge_map[(row, col)]
    return ws.cell(row, col).value


def parse_monthly_kpi_workbook(file_obj) -> ParsedKpiSheet:
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise KpiImportError(f'Không đọc được file Excel: {exc}') from exc

    ws = wb.active
    header_row = _find_header_row(ws)
    merge_map = _merged_top_left_map(ws)

    rows: list[ParsedKpiRow] = []
    errors: list[str] = []
    auto_stt = 0

    for r in range(header_row + 1, ws.max_row + 1):
        stt_raw = _cell_value(ws, r, 1, merge_map)
        group = _cell_str(_cell_value(ws, r, 2, merge_map))
        weight_raw = _cell_value(ws, r, 3, merge_map)
        indicator = _cell_str(_cell_value(ws, r, 4, merge_map))
        level_fail = _cell_str(_cell_value(ws, r, 5, merge_map))
        level_pass = _cell_str(_cell_value(ws, r, 6, merge_map))
        level_exceed = _cell_str(_cell_value(ws, r, 7, merge_map))

        # Bỏ dòng trống hoàn toàn
        if not any([
            _cell_str(stt_raw), group, _cell_str(weight_raw),
            indicator, level_fail, level_pass, level_exceed,
        ]):
            continue

        auto_stt += 1
        row_label = auto_stt
        missing = []
        if not _cell_str(stt_raw):
            missing.append('STT')
        if not group:
            missing.append('Nhóm công việc')
        if weight_raw is None or (isinstance(weight_raw, str) and not weight_raw.strip()):
            missing.append('Trọng số')
        if not indicator:
            missing.append('Tiêu chí đo lường')
        if not level_fail:
            missing.append('Mức Chưa đạt')
        if not level_pass:
            missing.append('Mức Đạt')
        if not level_exceed:
            missing.append('Mức Vượt')
        if missing:
            errors.append(f'Dòng Excel {r}: không được để trống — {", ".join(missing)}.')
            continue

        try:
            sort_order = int(float(stt_raw))
        except (TypeError, ValueError):
            errors.append(f'Dòng Excel {r}: STT không hợp lệ ({stt_raw!r}).')
            continue

        try:
            weightage = _to_float_required(weight_raw, row_no=r)
        except KpiImportError as exc:
            errors.append(str(exc))
            continue

        rows.append(ParsedKpiRow(
            sort_order=sort_order,
            work_group=group,
            weightage=weightage,
            indicator=indicator,
            level_fail=level_fail,
            level_pass=level_pass,
            level_exceed=level_exceed,
        ))

    if errors:
        raise KpiImportError(' '.join(errors[:8]) + (f' … (+{len(errors) - 8} lỗi)' if len(errors) > 8 else ''))

    if not rows:
        raise KpiImportError('File không có dòng tiêu chí KPI hợp lệ.')

    return ParsedKpiSheet(rows=rows)


def _apply_border_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = _THIN


def _merge_same_work_groups(ws, data_start: int, data_end: int, group_col: int = 2) -> None:
    """Merge cột Nhóm công việc khi các dòng liên tiếp cùng nhóm."""
    if data_end < data_start:
        return
    start = data_start
    while start <= data_end:
        group = _cell_str(ws.cell(start, group_col).value)
        end = start
        while end + 1 <= data_end and _cell_str(ws.cell(end + 1, group_col).value) == group:
            end += 1
        if end > start and group:
            ws.merge_cells(
                start_row=start,
                start_column=group_col,
                end_row=end,
                end_column=group_col,
            )
            ws.cell(start, group_col).alignment = Alignment(
                wrap_text=True, vertical='center', horizontal='center',
            )
        start = end + 1


def _add_scoring_guide_sheet(wb) -> None:
    ws = wb.create_sheet('Hướng dẫn chấm điểm')
    ws['A1'] = _SCORING_GUIDE_TITLE
    ws['A1'].font = Font(bold=True, size=14, color='B91C1C')
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')

    ws['A3'] = _SCORING_GUIDE_BODY.strip()
    ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['A3'].font = Font(size=11)

    ws.merge_cells('A3:D28')
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 28
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 320

    for r in range(1, 29):
        for c in range(1, 5):
            ws.cell(r, c).border = _THIN


def build_monthly_kpi_sample_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KPI'

    header_font = Font(bold=True)
    header_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

    for col, title in enumerate(_HEADERS, start=1):
        cell = ws.cell(1, col, title)
        cell.font = header_font
        cell.alignment = header_align

    # Nhóm giống nhau ghi đủ trên mỗi dòng trước khi merge
    sample = [
        (1, 'Nhóm A (50%)', 25, '1. Tiêu chí mẫu 1', 'Chưa đạt mô tả', 'Đạt mô tả', 'Vượt mô tả'),
        (2, 'Nhóm A (50%)', 25, '2. Tiêu chí mẫu 2', 'Chưa đạt mô tả', 'Đạt mô tả', 'Vượt mô tả'),
        (3, 'Nhóm B (50%)', 50, '3. Tiêu chí mẫu 3', 'Chưa đạt mô tả', 'Đạt mô tả', 'Vượt mô tả'),
    ]
    for i, row in enumerate(sample, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(i, col, val)
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    data_end = 1 + len(sample)
    _merge_same_work_groups(ws, 2, data_end, group_col=2)
    _apply_border_range(ws, 1, data_end, 1, len(_HEADERS))

    widths = [6, 22, 10, 42, 28, 28, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36

    _add_scoring_guide_sheet(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

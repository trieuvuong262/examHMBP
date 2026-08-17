"""Import / export Excel cho danh mục thiết lập IE (SxRefBase)."""

from __future__ import annotations

import io
from datetime import datetime

from django.db import transaction
from django.http import HttpResponse

from san_xuat.ie_models import (
    SxProcessStage,
    SxProductPart,
    SxSkillLevel,
    SxSmvSource,
    SxStitchClass,
)
from san_xuat.services.operation_master import (
    ImportResult,
    OperationMasterImportError,
    _int,
    _s,
    _write_sheet,
    _yesno,
)

SHEET_GUIDE = '00_HUONG_DAN'
SHEET_DATA = 'DANH_MUC'
HEADERS = ['MÃ', 'TÊN', 'THỨ TỰ', 'ĐANG DÙNG', 'GHI CHÚ']

UPPERCASE_CODE_KINDS = frozenset({'bac-ky-nang', 'khau-san-xuat'})

REF_CATALOG_IO = {
    'cum-chi-tiet': {
        'model': SxProductPart,
        'label': 'Cụm chi tiết',
        'filename_export': 'Cum_Chi_Tiet',
        'filename_template': 'Mau_Import_Cum_Chi_Tiet',
        'sample_row': ['THÂN', 'Thân áo', 10, 'Có', 'Ví dụ — xóa trước khi import'],
    },
    'bac-ky-nang': {
        'model': SxSkillLevel,
        'label': 'Bậc kỹ năng',
        'filename_export': 'Bac_Ky_Nang',
        'filename_template': 'Mau_Import_Bac_Ky_Nang',
        'sample_row': ['A', 'Bậc A', 10, 'Có', 'Ví dụ — xóa trước khi import'],
    },
    'khau-san-xuat': {
        'model': SxProcessStage,
        'label': 'Khâu sản xuất',
        'filename_export': 'Khau_San_Xuat',
        'filename_template': 'Mau_Import_Khau_San_Xuat',
        'sample_row': ['CAT', 'Cắt', 10, 'Có', 'Ví dụ — xóa trước khi import'],
    },
    'lop-mui-may': {
        'model': SxStitchClass,
        'label': 'Lớp mũi may',
        'filename_export': 'Lop_Mui_May',
        'filename_template': 'Mau_Import_Lop_Mui_May',
        'sample_row': ['301', 'Mũi 301', 10, 'Có', 'Ví dụ — xóa trước khi import'],
    },
    'nguon-smv': {
        'model': SxSmvSource,
        'label': 'Nguồn SMV',
        'filename_export': 'Nguon_SMV',
        'filename_template': 'Mau_Import_Nguon_SMV',
        'sample_row': ['TS', 'Time study', 10, 'Có', 'Ví dụ — xóa trước khi import'],
    },
}

REF_CATALOG_IO_KINDS = tuple(REF_CATALOG_IO.keys())


class RefCatalogImportError(OperationMasterImportError):
    pass


def normalize_ref_kind(kind: str | None) -> str:
    raw = (kind or '').strip().lower()
    if raw not in REF_CATALOG_IO:
        raise RefCatalogImportError(
            f'Danh mục không hỗ trợ import/export: {kind!r}. '
            f'Hợp lệ: {", ".join(REF_CATALOG_IO_KINDS)}.'
        )
    return raw


def ref_catalog_io_meta(kind: str | None) -> dict:
    kind = normalize_ref_kind(kind)
    cfg = REF_CATALOG_IO[kind]
    return {
        'kind': kind,
        'label': cfg['label'],
        'sheet': SHEET_DATA,
        'step_fill': f'Điền sheet {SHEET_DATA}',
        'drop_hint': f'.xlsx / .xlsm — giữ đúng tên sheet {SHEET_DATA}',
        'required_badges': [
            ('MÃ', True),
            ('TÊN', False),
            ('THỨ TỰ', False),
            ('ĐANG DÙNG', False),
            ('GHI CHÚ', False),
        ],
    }


def _normalize_code(kind: str, code: str) -> str:
    code = code.strip()[:40]
    if kind in UPPERCASE_CODE_KINDS:
        code = code.upper()
    return code


def _guide_rows(label: str) -> list[list]:
    return [
        [f'HƯỚNG DẪN NHẬP {label.upper()}'],
        [],
        ['Bước 1', 'Đọc sheet này — không đổi tên sheet dữ liệu.'],
        ['Bước 2', f'Điền sheet {SHEET_DATA}: mỗi dòng = 1 mục danh mục.'],
        ['Bước 3', 'Xóa dòng ví dụ mẫu, điền dữ liệu thật, lưu .xlsx.'],
        ['Bước 4', f'Portal → Thiết lập → {label} → Import → nên bật “Chạy thử” lần đầu.'],
        [],
        ['Cột bắt buộc', 'MÃ'],
        ['Cột tùy chọn', 'TÊN · THỨ TỰ · ĐANG DÙNG (Có/Không) · GHI CHÚ'],
        ['Trùng mã', 'Cùng MÃ → hệ thống CẬP NHẬT bản ghi cũ.'],
        ['Mẹo', 'Xuất Excel danh mục hiện tại rồi chỉnh — dễ hơn điền từ file trống.'],
    ]


def _export_rows(kind: str) -> list[list]:
    cfg = REF_CATALOG_IO[kind]
    Model = cfg['model']
    return [
        [
            obj.code,
            obj.name,
            obj.sort_order,
            'Có' if obj.is_active else 'Không',
            obj.notes or '',
        ]
        for obj in Model.objects.order_by('sort_order', 'code')
    ]


def export_ref_catalog_workbook(kind: str, *, template: bool = False):
    kind = normalize_ref_kind(kind)
    cfg = REF_CATALOG_IO[kind]
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise RefCatalogImportError('Thiếu thư viện openpyxl.') from exc

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_GUIDE
    for row in _guide_rows(cfg['label']):
        ws.append(row)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 88

    ws_data = wb.create_sheet(SHEET_DATA)
    rows = [cfg['sample_row']] if template else _export_rows(kind)
    _write_sheet(ws_data, HEADERS, rows)
    return wb


def export_ref_catalog_response(kind: str, *, template: bool = False, user=None):
    kind = normalize_ref_kind(kind)
    cfg = REF_CATALOG_IO[kind]
    wb = export_ref_catalog_workbook(kind, template=template)
    buf = io.BytesIO()
    wb.save(buf)
    stamp = datetime.now().strftime('%Y%m%d' if template else '%Y%m%d_%H%M')
    prefix = cfg['filename_template'] if template else cfg['filename_export']
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={prefix}_{stamp}.xlsx'
    if not template and user is not None:
        from san_xuat.ie_models import SxIeAuditLog
        from san_xuat.services.ie_audit import log_ie_event

        log_ie_event(
            action=SxIeAuditLog.ACTION_EXPORT,
            summary=f'Xuất Excel {cfg["label"]}',
            object_type=cfg['model'].__name__,
            object_repr=f'export_excel:{kind}',
            user=user,
        )
    return response


def _sheet_dicts(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_s(h) for h in rows[0]]
    out = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            item[h.casefold()] = row[i] if i < len(row) else None
        out.append(item)
    return out


def _row_value(row: dict, *keys: str):
    for key in keys:
        val = row.get(key.casefold())
        if val is not None and str(val).strip():
            return val
    return None


def import_ref_catalog(source, kind: str, *, dry_run: bool = False, user=None) -> ImportResult:
    kind = normalize_ref_kind(kind)
    cfg = REF_CATALOG_IO[kind]
    Model = cfg['model']
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RefCatalogImportError('Thiếu thư viện openpyxl.') from exc

    import warnings as _w

    result = ImportResult()
    try:
        with _w.catch_warnings():
            _w.simplefilter('ignore')
            wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    except Exception as exc:
        raise RefCatalogImportError(f'Không đọc được file Excel: {exc}') from exc

    try:
        if SHEET_DATA not in wb.sheetnames:
            raise RefCatalogImportError(
                f'Thiếu sheet {SHEET_DATA}. File này dành cho «{cfg["label"]}» '
                f'(1 sheet hướng dẫn + 1 sheet dữ liệu).'
            )
        rows = _sheet_dicts(wb[SHEET_DATA])
        if not rows:
            result.warnings.append(f'Sheet {SHEET_DATA} không có dòng dữ liệu.')
            return result

        with transaction.atomic():
            for line_no, row in enumerate(rows, start=2):
                raw_code = _s(_row_value(row, 'MÃ', 'MA', 'CODE'))
                if not raw_code:
                    result.warnings.append(f'Dòng {line_no}: thiếu MÃ — bỏ qua.')
                    continue
                code = _normalize_code(kind, raw_code)
                name_raw = _s(_row_value(row, 'TÊN', 'TEN', 'NAME'))
                name = name_raw[:150] if name_raw else code
                sort_raw = _row_value(row, 'THỨ TỰ', 'THU TU', 'SORT_ORDER', 'SORT')
                sort_order = _int(sort_raw) if sort_raw is not None and str(sort_raw).strip() else 100
                if sort_order is None:
                    sort_order = 100
                active_raw = _row_value(row, 'ĐANG DÙNG', 'DANG DUNG', 'HIỆU LỰC', 'HIEU LUC', 'IS_ACTIVE')
                is_active = _yesno(active_raw) if active_raw is not None and str(active_raw).strip() else True
                notes = _s(_row_value(row, 'GHI CHÚ', 'GHI CHU', 'NOTES'))[:255]

                _, created = Model.objects.update_or_create(
                    code=code,
                    defaults={
                        'name': name,
                        'sort_order': sort_order,
                        'is_active': is_active,
                        'notes': notes,
                    },
                )
                result.bump('row', created)

            if dry_run:
                transaction.set_rollback(True)
    finally:
        wb.close()

    if not dry_run and user is not None:
        from san_xuat.ie_models import SxIeAuditLog
        from san_xuat.services.ie_audit import log_ie_event

        log_ie_event(
            action=SxIeAuditLog.ACTION_IMPORT,
            summary=(
                f'Import {cfg["label"]} — tạo {result.total_created}, '
                f'cập nhật {result.total_updated}'
            ),
            object_type=Model.__name__,
            object_repr=f'import_excel:{kind}',
            changes={
                'kind': kind,
                'created': result.created,
                'updated': result.updated,
                'warnings': len(result.warnings),
            },
            user=user,
        )
    return result

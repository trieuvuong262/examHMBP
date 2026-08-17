"""Import Master Data Mã Công Đoạn Sản Xuất từ file Excel.

UI tách 3 file độc lập (mỗi file = 00_HUONG_DAN + 1 sheet dữ liệu):
  groups  → 01_DM_NHOM_CONG_DOAN
  library → 02_THU_VIEN_CONG_DOAN
  routing → 03_ROUTING_MA_HANG

CLI / master đầy đủ vẫn hỗ trợ import_operation_master (nhiều sheet):
  05_DM_THAM_CHIEU → danh mục nền
  01…04 → nhóm, thư viện, routing, time study

Idempotent: update_or_create theo khóa tự nhiên.
"""

from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from django.db import transaction

from san_xuat.ie_models import (
    SxMachine,
    SxOperation,
    SxOperationGroup,
    SxProcessStage,
    SxRouting,
    SxRoutingLine,
    SxSkillLevel,
    SxSmvBasis,
    SxSmvSource,
    SxStitchClass,
    SxTimeStudy,
    ensure_process_stage_defaults,
    ensure_skill_levels_abc,
    ensure_smv_basis_defaults,
    default_smv_basis_name,
    normalize_skill_level_label,
)
from san_xuat.hub_models import SxWorkCenter
from san_xuat.services.ie_ops import (
    link_time_studies_to_operations,
    operation_library_snapshot,
    resolve_operation,
)

# Tên sheet
SHEET_GUIDE = '00_HUONG_DAN'
SHEET_REF = '05_DM_THAM_CHIEU'
SHEET_GROUP = '01_DM_NHOM_CONG_DOAN'
SHEET_LIB = '02_THU_VIEN_CONG_DOAN'
SHEET_ROUTING = '03_ROUTING_MA_HANG'
SHEET_TIMESTUDY = '04_DU_LIEU_TIME_STUDY'
SHEET_DASHBOARD = '07_DASHBOARD'

# Dataset tách riêng (mỗi file = 1 sheet hướng dẫn + 1 sheet dữ liệu)
KIND_GROUPS = 'groups'
KIND_LIBRARY = 'library'
KIND_ROUTING = 'routing'
IE_DATASET_KINDS = (KIND_GROUPS, KIND_LIBRARY, KIND_ROUTING)

GROUP_HEADERS = [
    'MÃ NHÓM', 'TÊN NHÓM', 'KHÂU SẢN XUẤT', 'SẢN PHẨM CẦN', 'MÔ TẢ CHI TIẾT',
    'HIỆU LỰC', 'NGƯỜI LẬP', 'NGÀY HIỆU LỰC', 'NOTES',
]
LIB_HEADERS = [
    'MÃ NHÓM', 'MÃ CÔNG ĐOẠN', 'PHIÊN BẢN', 'TÊN CÔNG ĐOẠN', 'TÊN CÔNG ĐOẠN_EN',
    'BẬC CÔNG ĐOẠN', 'ĐỊNH MỨC THỜI GIAN', 'ĐỊNH MỨC SP/H', 'KHÂU SẢN XUẤT',
    'CỤM CHI TIẾT CHÍNH', 'MÔ TẢ PHƯƠNG PHÁP', 'MÃ MÁY MÓC', 'NHÓM MŨI MAY',
    'QUY ĐỊNH KIM/CHỈ', 'MÃ CỮ/GIÁ/CHÂN VỊT', 'ĐƠN VỊ', 'NGUỒN SMV', 'TRẠNG THÁI',
    'NGÀY HIỆU LỰC', 'NGÀY HẾT HIỆU LỰC', 'NGƯỜI LẬP', 'NGƯỜI DUYỆT',
    'LÝ DO CHỈNH SỬA PHIÊN BẢN', 'VIDEO_URL', 'NOTES',
]
ROUTING_HEADERS = [
    'MÃ ĐƠN HÀNG', 'MÃ HÀNG SẢN PHẨM', 'TÊN MÃ HÀNG', 'NHÓM SẢN PHẨM',
    'MÃ NHÓM', 'MÃ CÔNG ĐOẠN', 'PHIÊN BẢN', 'TÊN CÔNG ĐOẠN', 'SỐ LƯỢNG',
    'BẬC CÔNG ĐOẠN', 'ĐỊNH MỨC THỜI GIAN', 'ĐỊNH MỨC THEO PHIÊN BẢN',
    'TỔNG ĐỊNH MỨC', 'ĐỊNH MỨC SP/H', '% CHÊNH LỆCH ĐỊNH MỨC THỜI GIAN',
    'HỆ SỐ ĐƠN GIÁ', 'TỔNG ĐƠN GIÁ', 'TRẠNG THÁI ÁP DỤNG', 'NGÀY HIỆU LỰC',
    'MÃ MÁY MÓC', 'WORK_CENTER', 'NGƯỜI LẬP', 'NOTES',
]

VARIANCE_WARN_PCT = Decimal('15')

_STATUS_MAP = {
    'nháp': SxOperation.STATUS_DRAFT,
    'thử nghiệm': SxOperation.STATUS_TRIAL,
    'đã duyệt': SxOperation.STATUS_APPROVED,
    'ngưng sử dụng': SxOperation.STATUS_RETIRED,
}
_APPROVAL_MAP = {
    'chờ duyệt': SxTimeStudy.APPROVAL_PENDING,
    'đã duyệt': SxTimeStudy.APPROVAL_APPROVED,
    'từ chối': SxTimeStudy.APPROVAL_REJECTED,
    'cần đo lại': SxTimeStudy.APPROVAL_REMEASURE,
}


@dataclass
class ImportResult:
    created: dict[str, int] = field(default_factory=dict)
    updated: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def bump(self, key: str, created: bool) -> None:
        target = self.created if created else self.updated
        target[key] = target.get(key, 0) + 1

    @property
    def total_created(self) -> int:
        return sum(self.created.values())

    @property
    def total_updated(self) -> int:
        return sum(self.updated.values())


class OperationMasterImportError(Exception):
    pass


# --- Helpers ---------------------------------------------------------------


def _s(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _dec(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    return None


def _int(value):
    d = _dec(value)
    if d is None:
        return None
    return int(d)


def _date(value):
    if value is None or value == '':
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    txt = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return _dt.datetime.strptime(txt[:10], fmt).date()
        except ValueError:
            continue
    return None


def _yesno(value) -> bool:
    return _s(value).casefold() in ('có', 'co', 'yes', 'true', '1', 'x')


def _sheet_dicts(ws):
    """Đọc sheet dạng bảng: dòng đầu là header, trả về list dict theo header (đã strip)."""
    rows = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return []
    headers = [(_s(h)) for h in raw_headers]
    out = []
    for row in rows:
        rec = {}
        has_value = False
        for i, head in enumerate(headers):
            if not head:
                continue
            val = row[i] if i < len(row) else None
            rec[head] = val
            if val not in (None, ''):
                has_value = True
        if has_value:
            out.append(rec)
    return out


def _col_values(ws, header_name):
    """Lấy các giá trị không rỗng của một cột theo tên header (sheet dạng nhiều cột danh mục)."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_s(h) for h in rows[0]]
    try:
        idx = headers.index(header_name)
    except ValueError:
        return []
    values = []
    for row in rows[1:]:
        if idx < len(row):
            v = _s(row[idx])
            if v:
                values.append(v)
    return values


def _col_pairs(ws, header_a, header_b):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_s(h) for h in rows[0]]
    try:
        ia = headers.index(header_a)
        ib = headers.index(header_b)
    except ValueError:
        return []
    pairs = []
    for row in rows[1:]:
        a = _s(row[ia]) if ia < len(row) else ''
        b = _s(row[ib]) if ib < len(row) else ''
        if a:
            pairs.append((a, b))
    return pairs


# --- Import từng sheet -----------------------------------------------------


def _is_ie_template_sample_row(rec: dict) -> bool:
    """True nếu dòng Excel là dòng ví dụ trong file mẫu — không import vào DB."""
    notes = _s(rec.get('NOTES')).casefold()
    if 'dòng mẫu' in notes or 'dong mau' in notes:
        return True
    name = _s(rec.get('TÊN CÔNG ĐOẠN')).casefold()
    name_en = _s(rec.get('TÊN CÔNG ĐOẠN_EN')).casefold()
    if '(ví dụ)' in name or '(vi du)' in name or '(sample)' in name or '(sample)' in name_en:
        return True
    group_desc = _s(rec.get('MÔ TẢ CHI TIẾT')).casefold()
    if 'nhóm ví dụ' in group_desc or 'nhom vi du' in group_desc:
        return True
    style_name = _s(rec.get('TÊN MÃ HÀNG')).casefold()
    if 'mã hàng ví dụ' in style_name or 'ma hang vi du' in style_name:
        return True
    routing_id = _s(rec.get('MÃ ĐƠN HÀNG')).casefold()
    style_code = _s(rec.get('MÃ HÀNG SẢN PHẨM')).casefold()
    if routing_id.startswith('style-demo') or style_code == 'style-demo':
        return True
    # Mã mẫu cố định trong file template thư viện
    op_code = _s(rec.get('MÃ CÔNG ĐOẠN')).upper()
    if op_code in {'SEW-1001', 'SEW-1002'} and (
        '(ví dụ)' in _s(rec.get('TÊN CÔNG ĐOẠN')).casefold()
        or 'sample' in name_en
        or 'dòng mẫu' in notes
    ):
        return True
    return False


def _import_reference(wb, result: ImportResult) -> None:
    if SHEET_REF not in wb.sheetnames:
        result.warnings.append(f'Không thấy sheet {SHEET_REF}, bỏ qua danh mục nền.')
        return
    ws = wb[SHEET_REF]

    for order, (code, name) in enumerate(_col_pairs(ws, 'MACHINE_CODE', 'MACHINE_NAME'), start=1):
        _, created = SxMachine.objects.update_or_create(
            code=code, defaults={'name': name or code, 'sort_order': order * 10},
        )
        result.bump('machine', created)

    simple = [
        ('SKILL_LEVEL', SxSkillLevel, 'skill_level'),
        ('SMV_SOURCE', SxSmvSource, 'smv_source'),
        ('SMV_BASIS', SxSmvBasis, 'smv_basis'),
        ('PROCESS_STAGE', SxProcessStage, 'process_stage'),
        ('STITCH_CLASS', SxStitchClass, 'stitch_class'),
    ]
    for header, model, key in simple:
        for order, value in enumerate(_col_values(ws, header), start=1):
            if model is SxSkillLevel:
                value = normalize_skill_level_label(value)
                if value not in ('A', 'B', 'C'):
                    continue
            _, created = model.objects.update_or_create(
                code=value, defaults={'name': value, 'sort_order': order * 10},
            )
            result.bump(key, created)

    ensure_skill_levels_abc()
    ensure_process_stage_defaults()
    ensure_smv_basis_defaults()

    for order, code in enumerate(_col_values(ws, 'WORK_CENTER'), start=1):
        # Không tạo thêm WC IE — bộ phận chuẩn lấy từ HR (HRD-*). Chỉ bỏ qua mã lạ.
        from san_xuat.services.capacity_from_hrm import resolve_work_center_code
        if resolve_work_center_code(code):
            result.bump('work_center', False)
        else:
            result.warnings.append(
                f'[Ref] WORK_CENTER {code} không map được sang bộ phận HR — bỏ qua (dùng Đồng bộ HR).'
            )


def _import_groups(wb, result: ImportResult) -> None:
    if SHEET_GROUP not in wb.sheetnames:
        result.warnings.append(f'Không thấy sheet {SHEET_GROUP}, bỏ qua nhóm công đoạn.')
        return
    from san_xuat.services.capacity_from_hrm import resolve_work_center_code

    seen = set()
    skipped_samples = 0
    for order, rec in enumerate(_sheet_dicts(wb[SHEET_GROUP]), start=1):
        code = _s(rec.get('MÃ NHÓM'))
        if not code:
            continue
        if _is_ie_template_sample_row(rec):
            skipped_samples += 1
            continue
        if code in seen:
            result.warnings.append(f'[Nhóm] Trùng MÃ NHÓM: {code}')
            continue
        seen.add(code)
        stage_label = _s(rec.get('KHÂU SẢN XUẤT'))
        wc_code = _s(rec.get('DEFAULT_WORK_CENTER'))  # tùy chọn (file cũ)
        stage = SxProcessStage.objects.filter(name=stage_label).first() if stage_label else None
        wc = resolve_work_center_code(wc_code, name_hint=f'{stage_label} {_s(rec.get("TÊN NHÓM"))}')
        product_part = _s(rec.get('SẢN PHẨM CẦN')) or _s(rec.get('SẢN PHẨM CÂN'))
        _, created = SxOperationGroup.objects.update_or_create(
            code=code,
            defaults={
                'name': _s(rec.get('TÊN NHÓM')),
                'process_stage': stage,
                'process_stage_label': stage_label,
                'product_part': product_part,
                'description': _s(rec.get('MÔ TẢ CHI TIẾT')),
                'default_work_center': wc,
                'default_work_center_code': wc.code if wc else '',
                'data_owner': _s(rec.get('NGƯỜI LẬP')),
                'effective_from': _date(rec.get('NGÀY HIỆU LỰC')),
                'is_active': _yesno(rec.get('HIỆU LỰC')) if rec.get('HIỆU LỰC') is not None else True,
                'sort_order': order * 10,
                'notes': _s(rec.get('NOTES')),
            },
        )
        result.bump('group', created)
        if wc_code and not wc:
            result.warnings.append(f'[Nhóm] {code}: DEFAULT_WORK_CENTER {wc_code} không map được sang bộ phận HR.')
    if skipped_samples:
        result.warnings.append(
            f'[Nhóm] Đã bỏ qua {skipped_samples} dòng ví dụ mẫu (không import vào hệ thống).'
        )


def _import_operations(wb, result: ImportResult, *, user=None) -> None:
    if SHEET_LIB not in wb.sheetnames:
        result.warnings.append(f'Không thấy sheet {SHEET_LIB}, bỏ qua thư viện công đoạn.')
        return
    from san_xuat.ie_permissions import ie_user_display_name

    importer_name = ie_user_display_name(user)
    seen = set()
    skipped_samples = 0
    for rec in _sheet_dicts(wb[SHEET_LIB]):
        op_code = _s(rec.get('MÃ CÔNG ĐOẠN'))
        if not op_code:
            continue
        if _is_ie_template_sample_row(rec):
            skipped_samples += 1
            continue
        op_rev = _s(rec.get('PHIÊN BẢN')) or 'R01'
        key = (op_code, op_rev)
        if key in seen:
            result.warnings.append(f'[Công đoạn] Trùng OP_CODE+OP_REV: {op_code}/{op_rev}')
            continue
        seen.add(key)

        group_code = _s(rec.get('MÃ NHÓM'))
        group = SxOperationGroup.objects.filter(code=group_code).first()
        if group is None:
            group, _ = SxOperationGroup.objects.get_or_create(
                code=group_code or f'AUTO-{op_code}',
                defaults={'name': group_code or op_code},
            )
            result.warnings.append(f'[Công đoạn] {op_code}: tạo nhóm tạm {group.code} (thiếu trong sheet nhóm).')

        time_sec = _dec(rec.get('ĐỊNH MỨC THỜI GIAN')) or Decimal('0')
        base_smv = time_sec.quantize(Decimal('0.0001'))
        if base_smv <= 0:
            result.warnings.append(f'[Công đoạn] {op_code}/{op_rev}: SMV chuẩn = 0.')

        machine_code = _s(rec.get('MÃ MÁY MÓC'))
        stitch_val = _s(rec.get('NHÓM MŨI MAY'))
        skill_label = normalize_skill_level_label(_s(rec.get('BẬC CÔNG ĐOẠN')))
        smv_source_label = _s(rec.get('NGUỒN SMV'))
        status_label = _s(rec.get('TRẠNG THÁI')).casefold()
        smv_basis = default_smv_basis_name()
        ie_owner = importer_name or _s(rec.get('NGƯỜI LẬP'))

        _, created = SxOperation.objects.update_or_create(
            op_code=op_code,
            op_rev=op_rev,
            defaults={
                'group': group,
                'name_vi': _s(rec.get('TÊN CÔNG ĐOẠN')),
                'name_en': _s(rec.get('TÊN CÔNG ĐOẠN_EN')),
                'process_stage_label': _s(rec.get('KHÂU SẢN XUẤT')),
                'product_part': _s(rec.get('CỤM CHI TIẾT CHÍNH')),
                'method_variant': _s(rec.get('MÔ TẢ PHƯƠNG PHÁP')),
                'machine': SxMachine.objects.filter(code=machine_code).first() if machine_code else None,
                'machine_code': machine_code,
                'stitch_class': SxStitchClass.objects.filter(code=stitch_val).first() if stitch_val else None,
                'thread_needle': _s(rec.get('QUY ĐỊNH KIM/CHỈ')),
                'attachment_code': _s(rec.get('MÃ CỮ/GIÁ/CHÂN VỊT')),
                'smv_basis': smv_basis,
                'skill_level': (
                    SxSkillLevel.objects.filter(code=skill_label).first()
                    or SxSkillLevel.objects.filter(name=skill_label).first()
                ) if skill_label else None,
                'skill_level_label': skill_label,
                'base_smv_min': base_smv,
                'smv_source': SxSmvSource.objects.filter(name=smv_source_label).first() if smv_source_label else None,
                'status': _STATUS_MAP.get(status_label, SxOperation.STATUS_DRAFT),
                'effective_from': _date(rec.get('NGÀY HIỆU LỰC')),
                'effective_to': _date(rec.get('NGÀY HẾT HIỆU LỰC')),
                'ie_owner': ie_owner,
                'approved_by': _s(rec.get('NGƯỜI DUYỆT')),
                'revision_reason': _s(rec.get('LÝ DO CHỈNH SỬA PHIÊN BẢN')),
                'video_url': _s(rec.get('VIDEO_URL')),
                'notes': _s(rec.get('NOTES')),
            },
        )
        result.bump('operation', created)
    if skipped_samples:
        result.warnings.append(
            f'[Công đoạn] Đã bỏ qua {skipped_samples} dòng ví dụ mẫu (không import vào hệ thống).'
        )


def _routing_rev_from_id(routing_id: str, fallback: str) -> str:
    parts = routing_id.rsplit('-', 1)
    if len(parts) == 2 and parts[1][:1].upper() == 'R':
        return parts[1]
    return fallback or 'R01'


def _import_routings(wb, result: ImportResult, *, user=None) -> None:
    if SHEET_ROUTING not in wb.sheetnames:
        result.warnings.append(f'Không thấy sheet {SHEET_ROUTING}, bỏ qua routing.')
        return

    from san_xuat.ie_permissions import ie_user_display_name

    importer_name = ie_user_display_name(user)
    grouped: dict[str, list[dict]] = {}
    order_ids: list[str] = []
    skipped_samples = 0
    for rec in _sheet_dicts(wb[SHEET_ROUTING]):
        routing_id = _s(rec.get('MÃ ĐƠN HÀNG'))
        op_code = _s(rec.get('MÃ CÔNG ĐOẠN'))
        if not routing_id or not op_code:
            continue
        if _is_ie_template_sample_row(rec):
            skipped_samples += 1
            continue
        if routing_id not in grouped:
            grouped[routing_id] = []
            order_ids.append(routing_id)
        grouped[routing_id].append(rec)

    for routing_id in order_ids:
        recs = grouped[routing_id]
        head = recs[0]
        existing = SxRouting.objects.filter(routing_id=routing_id).first()
        if existing is not None:
            from san_xuat.hub_models import SxProductionOrder

            if SxProductionOrder.objects.filter(routing_id=existing.pk).exists():
                result.warnings.append(
                    f'[Routing] {routing_id}: đã gắn lệnh SX — bỏ qua import (không sửa đè).'
                )
                continue

        routing, created = SxRouting.objects.update_or_create(
            routing_id=routing_id,
            defaults={
                'style_code': _s(head.get('MÃ HÀNG SẢN PHẨM')),
                'style_name': _s(head.get('TÊN MÃ HÀNG')),
                'product_family': _s(head.get('NHÓM SẢN PHẨM')),
                'routing_rev': _routing_rev_from_id(routing_id, _s(head.get('PHIÊN BẢN'))),
                'effective_from': _date(head.get('NGÀY HIỆU LỰC')),
                'is_active': _yesno(head.get('TRẠNG THÁI ÁP DỤNG')) if head.get('TRẠNG THÁI ÁP DỤNG') is not None else True,
                'ie_owner': importer_name or _s(head.get('NGƯỜI LẬP')),
                'notes': _s(head.get('NOTES')),
                'approval_status': SxRouting.APPROVAL_APPROVED,
            },
        )
        result.bump('routing', created)

        # Làm mới các dòng cho routing này (tránh sót/trùng SEQ).
        routing.lines.all().delete()
        seq = 0
        for rec in recs:
            seq += 10
            op_code = _s(rec.get('MÃ CÔNG ĐOẠN'))
            op_rev = _s(rec.get('PHIÊN BẢN')) or 'R01'
            op = resolve_operation(op_code, op_rev)
            machine_code = _s(rec.get('MÃ MÁY MÓC'))
            op_name = _s(rec.get('TÊN CÔNG ĐOẠN'))
            group_code = _s(rec.get('MÃ NHÓM'))
            if op:
                snap = operation_library_snapshot(op)
                if not op_name:
                    op_name = snap.get('name_vi', '')
                if not group_code:
                    group_code = snap.get('group_code', '')
                if not machine_code:
                    machine_code = snap.get('machine_code', '')
            wc_code = _s(rec.get('WORK_CENTER'))
            applied = _dec(rec.get('ĐỊNH MỨC THỜI GIAN')) or Decimal('0')
            library = _dec(rec.get('ĐỊNH MỨC THEO PHIÊN BẢN')) or Decimal('0')
            if not library and op:
                library = op.base_smv_min or Decimal('0')
            qty = _dec(rec.get('SỐ LƯỢNG')) or _dec(rec.get('SỐ LƯỢNG ')) or Decimal('1')
            variance_text = ''
            if library and applied and library > 0:
                pct = abs((applied - library) / library * Decimal('100'))
                if pct > VARIANCE_WARN_PCT:
                    variance_text = f'Import Excel — lệch {pct.quantize(Decimal("0.01"))}%'
            from san_xuat.services.capacity_from_hrm import resolve_work_center_code
            wc = resolve_work_center_code(
                wc_code,
                name_hint=f'{_s(rec.get("MÃ NHÓM"))} {_s(rec.get("TÊN CÔNG ĐOẠN"))}',
            )
            line = SxRoutingLine(
                routing=routing,
                seq_no=seq,
                operation=op,
                op_code=op_code,
                op_rev=op_rev,
                op_name_vi=op_name,
                group_code=group_code,
                qty_per_garment=qty,
                library_unit_smv=library,
                applied_unit_smv=applied,
                price_factor=_dec(rec.get('HỆ SỐ ĐƠN GIÁ')) or Decimal('0'),
                total_unit_price=_dec(rec.get('TỔNG ĐƠN GIÁ')) or Decimal('0'),
                machine=SxMachine.objects.filter(code=machine_code).first() if machine_code else None,
                machine_code=machine_code,
                work_center=wc,
                work_center_code=wc.code if wc else '',
                skill_level_label=normalize_skill_level_label(_s(rec.get('BẬC CÔNG ĐOẠN'))),
                notes=_s(rec.get('NOTES')),
                variance_explanation=variance_text[:500],
            )
            line.save()
            result.bump('routing_line', True)
            if applied <= 0:
                result.warnings.append(f'[Routing] {routing_id} #{seq} {op_code}: SMV áp dụng = 0.')
            elif abs(line.smv_variance_pct) > VARIANCE_WARN_PCT:
                result.warnings.append(
                    f'[Routing] {routing_id} #{seq} {op_code}: chênh lệch SMV {line.smv_variance_pct}% > {VARIANCE_WARN_PCT}%.'
                )
    if skipped_samples:
        result.warnings.append(
            f'[Routing] Đã bỏ qua {skipped_samples} dòng ví dụ mẫu (không import vào hệ thống).'
        )


def _import_time_studies(wb, result: ImportResult) -> None:
    if SHEET_TIMESTUDY not in wb.sheetnames:
        result.warnings.append(f'Không thấy sheet {SHEET_TIMESTUDY}, bỏ qua time study.')
        return
    for rec in _sheet_dicts(wb[SHEET_TIMESTUDY]):
        study_id = _s(rec.get('STUDY_ID'))
        if not study_id:
            continue
        op_code = _s(rec.get('OP_CODE'))
        op_rev = _s(rec.get('OP_REV')) or 'R01'
        machine_code = _s(rec.get('MACHINE_CODE'))
        approval_label = _s(rec.get('APPROVAL_STATUS')).casefold()
        _, created = SxTimeStudy.objects.update_or_create(
            study_id=study_id,
            defaults={
                'study_date': _date(rec.get('STUDY_DATE')),
                'factory_code': _s(rec.get('FACTORY_CODE')),
                'line_code': _s(rec.get('LINE_CODE')),
                'shift': _s(rec.get('SHIFT')),
                'style_code': _s(rec.get('STYLE_CODE')),
                'routing_rev': _s(rec.get('ROUTING_REV')),
                'operation': resolve_operation(op_code, op_rev),
                'op_code': op_code,
                'op_rev': op_rev,
                'op_name_vi': _s(rec.get('OP_NAME_VI')),
                'operator_id': _s(rec.get('OPERATOR_ID')),
                'skill_level_label': _s(rec.get('SKILL_LEVEL')),
                'machine': SxMachine.objects.filter(code=machine_code).first() if machine_code else None,
                'machine_code': machine_code,
                'method_rev': _s(rec.get('METHOD_REV')),
                'obs_no': _int(rec.get('OBS_NO')) or 1,
                'observed_cycle_sec': _dec(rec.get('OBSERVED_CYCLE_SEC')) or Decimal('0'),
                'abnormal_sec': _dec(rec.get('ABNORMAL_SEC')) or Decimal('0'),
                'performance_rating': _dec(rec.get('PERFORMANCE_RATING')) or Decimal('0'),
                'allowance_pct': _dec(rec.get('ALLOWANCE_PCT')) or Decimal('0'),
                'current_routing_smv': _dec(rec.get('CURRENT_ROUTING_SMV')) or Decimal('0'),
                'ie_observer': _s(rec.get('IE_OBSERVER')),
                'conditions': _s(rec.get('CONDITIONS')),
                'approval_status': _APPROVAL_MAP.get(approval_label, SxTimeStudy.APPROVAL_PENDING),
                'notes': _s(rec.get('NOTES')),
            },
        )
        result.bump('time_study', created)


# --- Entry point -----------------------------------------------------------


def import_operation_master(source, *, dry_run: bool = False, user=None) -> ImportResult:
    """Import toàn bộ master data mã công đoạn từ file Excel.

    ``source`` là đường dẫn file hoặc file-like (upload). ``dry_run=True`` sẽ
    rollback sau khi chạy (chỉ để kiểm tra & thống kê).
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise OperationMasterImportError('Thiếu thư viện openpyxl.') from exc

    import warnings as _w

    result = ImportResult()
    try:
        with _w.catch_warnings():
            _w.simplefilter('ignore')
            wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    except Exception as exc:
        raise OperationMasterImportError(f'Không đọc được file Excel: {exc}') from exc

    try:
        with transaction.atomic():
            _import_reference(wb, result)
            _import_groups(wb, result)
            _import_operations(wb, result, user=user)
            _import_routings(wb, result, user=user)
            _import_time_studies(wb, result)
            link_stats = link_time_studies_to_operations(only_unlinked=True)
            if link_stats['linked']:
                result.warnings.append(
                    f"Đã gắn FK time study → operation: {link_stats['linked']} quan sát."
                )
            if dry_run:
                transaction.set_rollback(True)
    finally:
        wb.close()

    if not dry_run:
        from san_xuat.ie_models import SxIeAuditLog
        from san_xuat.services.ie_audit import log_ie_event

        log_ie_event(
            action=SxIeAuditLog.ACTION_IMPORT,
            summary=(
                f'Import master data — tạo {result.total_created}, '
                f'cập nhật {result.total_updated}'
            ),
            object_type='OperationMaster',
            object_repr='import_excel',
            changes={
                'created': result.created,
                'updated': result.updated,
                'warnings': len(result.warnings),
            },
            user=user,
        )
    return result


def normalize_ie_kind(kind: str | None) -> str:
    raw = (kind or '').strip().lower()
    aliases = {
        'group': KIND_GROUPS,
        'nhom': KIND_GROUPS,
        'groups': KIND_GROUPS,
        'lib': KIND_LIBRARY,
        'ops': KIND_LIBRARY,
        'operation': KIND_LIBRARY,
        'operations': KIND_LIBRARY,
        'thu-vien': KIND_LIBRARY,
        'library': KIND_LIBRARY,
        'route': KIND_ROUTING,
        'routes': KIND_ROUTING,
        'routing': KIND_ROUTING,
        'ma-hang': KIND_ROUTING,
    }
    resolved = aliases.get(raw, raw)
    if resolved not in IE_DATASET_KINDS:
        raise OperationMasterImportError(
            f'Loại dữ liệu không hợp lệ: {kind!r}. Hợp lệ: {", ".join(IE_DATASET_KINDS)}.'
        )
    return resolved


def import_ie_dataset(source, kind, *, dry_run: bool = False, user=None) -> ImportResult:
    """Import một nhóm dữ liệu IE (nhóm / thư viện / routing) từ file Excel tách riêng."""
    kind = normalize_ie_kind(kind)
    meta = IE_DATASETS[kind]
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise OperationMasterImportError('Thiếu thư viện openpyxl.') from exc

    import warnings as _w

    result = ImportResult()
    try:
        with _w.catch_warnings():
            _w.simplefilter('ignore')
            wb = openpyxl.load_workbook(source, data_only=True, read_only=True)
    except Exception as exc:
        raise OperationMasterImportError(f'Không đọc được file Excel: {exc}') from exc

    try:
        if meta['sheet'] not in wb.sheetnames:
            raise OperationMasterImportError(
                f'Thiếu sheet {meta["sheet"]}. File này dành cho «{meta["label"]}» '
                f'(1 sheet hướng dẫn + 1 sheet dữ liệu).'
            )
        with transaction.atomic():
            if kind == KIND_GROUPS:
                _import_groups(wb, result)
            elif kind == KIND_LIBRARY:
                _import_operations(wb, result, user=user)
            else:
                _import_routings(wb, result, user=user)
            if dry_run:
                transaction.set_rollback(True)
    finally:
        wb.close()

    if not dry_run:
        from san_xuat.ie_models import SxIeAuditLog
        from san_xuat.services.ie_audit import log_ie_event

        log_ie_event(
            action=SxIeAuditLog.ACTION_IMPORT,
            summary=(
                f'Import {meta["label"]} — tạo {result.total_created}, '
                f'cập nhật {result.total_updated}'
            ),
            object_type='OperationMaster',
            object_repr=f'import_excel:{kind}',
            changes={
                'kind': kind,
                'created': result.created,
                'updated': result.updated,
                'warnings': len(result.warnings),
            },
            user=user,
        )
    return result


# --- Export ------------------------------------------------------------------

_STATUS_LABEL = {v: k.title() if k != 'đã duyệt' else 'Đã duyệt' for k, v in _STATUS_MAP.items()}
_STATUS_LABEL.update({
    SxOperation.STATUS_DRAFT: 'Nháp',
    SxOperation.STATUS_TRIAL: 'Thử nghiệm',
    SxOperation.STATUS_APPROVED: 'Đã duyệt',
    SxOperation.STATUS_RETIRED: 'Ngưng sử dụng',
})
_APPROVAL_LABEL = {
    SxTimeStudy.APPROVAL_PENDING: 'Chờ duyệt',
    SxTimeStudy.APPROVAL_APPROVED: 'Đã duyệt',
    SxTimeStudy.APPROVAL_REJECTED: 'Từ chối',
    SxTimeStudy.APPROVAL_REMEASURE: 'Cần đo lại',
}


def _write_sheet(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for row in rows:
        ws.append(row)


def _group_export_rows() -> list[list]:
    return [
        [
            g.code, g.name, g.process_stage_label, g.product_part, g.description,
            'Có' if g.is_active else 'Không', g.data_owner,
            g.effective_from.isoformat() if g.effective_from else '',
            g.notes,
        ]
        for g in SxOperationGroup.objects.order_by('sort_order', 'code')
    ]


def _library_export_rows() -> list[list]:
    lib_rows = []
    for op in SxOperation.objects.select_related('group').order_by('op_code', 'op_rev'):
        smv = op.base_smv_min or Decimal('0')
        time_sec = smv.quantize(Decimal('0.01'))
        pcs_h = (Decimal('3600') / smv).quantize(Decimal('0.01')) if smv else Decimal('0')
        lib_rows.append([
            op.group.code if op.group_id else '',
            op.op_code, op.op_rev, op.name_vi, op.name_en,
            op.skill_level_label, float(time_sec), float(pcs_h),
            op.process_stage_label, op.product_part, op.method_variant,
            op.machine_code,
            op.stitch_class.code if op.stitch_class_id else '',
            op.thread_needle, op.attachment_code, op.smv_basis,
            op.smv_source.name if op.smv_source_id else '',
            _STATUS_LABEL.get(op.status, op.status),
            op.effective_from.isoformat() if op.effective_from else '',
            op.effective_to.isoformat() if op.effective_to else '',
            op.ie_owner, op.approved_by, op.revision_reason,
            op.video_url, op.notes,
        ])
    return lib_rows


def _routing_export_rows() -> list[list]:
    route_rows = []
    for line in (
        SxRoutingLine.objects.select_related('routing')
        .order_by('routing__style_code', 'routing__routing_rev', 'seq_no')
    ):
        r = line.routing
        applied = line.applied_unit_smv or Decimal('0')
        pcs_h = (Decimal('3600') / applied).quantize(Decimal('0.01')) if applied else Decimal('0')
        route_rows.append([
            r.routing_id, r.style_code, r.style_name, r.product_family,
            line.group_code, line.op_code, line.op_rev, line.op_name_vi,
            float(line.qty_per_garment or 0), line.skill_level_label,
            float(applied), float(line.library_unit_smv or 0),
            float(line.total_operation_smv or 0), float(pcs_h),
            float(line.smv_variance_pct or 0),
            float(line.price_factor or 0), float(line.total_unit_price or 0),
            'Có' if r.is_active else 'Không',
            r.effective_from.isoformat() if r.effective_from else '',
            line.machine_code, line.work_center_code, r.ie_owner, line.notes,
        ])
    return route_rows


IE_DATASETS = {
    KIND_GROUPS: {
        'kind': KIND_GROUPS,
        'label': 'Nhóm công đoạn',
        'sheet': SHEET_GROUP,
        'headers': GROUP_HEADERS,
        'filename_export': 'Nhom_Cong_Doan',
        'filename_template': 'Mau_Import_Nhom_Cong_Doan',
        'step_fill': f'Điền sheet {SHEET_GROUP}',
        'drop_hint': f'.xlsx / .xlsm — giữ đúng tên sheet {SHEET_GROUP}',
        'required_badges': [
            ('MÃ NHÓM', True),
            ('TÊN NHÓM', True),
            ('KHÂU SẢN XUẤT', False),
            ('SẢN PHẨM CẦN', False),
            ('MÔ TẢ CHI TIẾT', False),
            ('HIỆU LỰC', False),
            ('NGƯỜI LẬP', False),
            ('NGÀY HIỆU LỰC', False),
            ('NOTES', False),
        ],
        'guide_rows': [
            ['HƯỚNG DẪN NHẬP NHÓM CÔNG ĐOẠN'],
            [],
            ['Bước 1', 'Đọc sheet này — không đổi tên sheet dữ liệu.'],
            ['Bước 2', f'Điền sheet {SHEET_GROUP}: mỗi dòng = 1 nhóm (Cắt, May, In-Ép…).'],
            ['Bước 3', 'Điền dữ liệu thật vào sheet dữ liệu (file mẫu chỉ có tiêu đề cột, không có dòng ví dụ).'],
            ['Bước 4', 'Portal → Nhóm công đoạn → Import.'],
            [],
            ['Cột bắt buộc', 'MÃ NHÓM · TÊN NHÓM'],
            ['Cột chi tiết', 'KHÂU SẢN XUẤT · SẢN PHẨM CẦN · MÔ TẢ CHI TIẾT'],
            ['Cột hiệu lực', 'HIỆU LỰC (Có/Không) · NGƯỜI LẬP · NGÀY HIỆU LỰC · NOTES'],
            ['Trùng mã', 'Cùng MÃ NHÓM → hệ thống CẬP NHẬT nhóm cũ.'],
            ['Mẹo', 'Xuất Excel nhóm hiện tại rồi chỉnh — dễ hơn điền từ file trống.'],
            ['Dòng mẫu cũ', 'Nếu file còn dòng NOTES «Dòng mẫu» / mô tả «ví dụ», hệ thống sẽ bỏ qua khi import.'],
        ],
        'sample_rows': [],
        'row_builder': _group_export_rows,
    },
    KIND_LIBRARY: {
        'kind': KIND_LIBRARY,
        'label': 'Thư viện công đoạn',
        'sheet': SHEET_LIB,
        'headers': LIB_HEADERS,
        'filename_export': 'Thu_Vien_Cong_Doan',
        'filename_template': 'Mau_Import_Thu_Vien_Cong_Doan',
        'step_fill': f'Điền sheet {SHEET_LIB}',
        'drop_hint': f'.xlsx / .xlsm — giữ đúng tên sheet {SHEET_LIB}',
        'required_badges': [
            ('MÃ NHÓM', False),
            ('MÃ CÔNG ĐOẠN', True),
            ('PHIÊN BẢN', True),
            ('TÊN CÔNG ĐOẠN', True),
            ('TÊN CÔNG ĐOẠN_EN', False),
            ('BẬC CÔNG ĐOẠN', False),
            ('ĐỊNH MỨC THỜI GIAN (giây)', False),
            ('ĐỊNH MỨC SP/H', False),
            ('KHÂU SẢN XUẤT', False),
            ('CỤM CHI TIẾT CHÍNH', False),
            ('MÔ TẢ PHƯƠNG PHÁP', False),
            ('MÃ MÁY MÓC', False),
            ('NHÓM MŨI MAY', False),
            ('QUY ĐỊNH KIM/CHỈ', False),
            ('MÃ CỮ/GIÁ/CHÂN VỊT', False),
            ('ĐƠN VỊ', False),
            ('NGUỒN SMV', False),
            ('TRẠNG THÁI', False),
            ('NGÀY HIỆU LỰC', False),
            ('NGÀY HẾT HIỆU LỰC', False),
            ('NGƯỜI LẬP', False),
            ('NGƯỜI DUYỆT', False),
            ('LÝ DO CHỈNH SỬA PHIÊN BẢN', False),
            ('VIDEO_URL', False),
            ('NOTES', False),
        ],
        'guide_rows': [
            ['HƯỚNG DẪN NHẬP THƯ VIỆN CÔNG ĐOẠN'],
            [],
            ['Bước 1', 'Đọc sheet này — không đổi tên sheet dữ liệu.'],
            ['Bước 2', f'Điền sheet {SHEET_LIB}: mỗi dòng = 1 công đoạn chuẩn (mã + phiên bản).'],
            ['Bước 3', 'Điền dữ liệu thật vào sheet dữ liệu (file mẫu chỉ có tiêu đề cột, không có dòng ví dụ).'],
            ['Bước 4', 'Portal → Thư viện công đoạn → Import.'],
            [],
            ['Cột bắt buộc', 'MÃ CÔNG ĐOẠN · TÊN CÔNG ĐOẠN · PHIÊN BẢN (mặc định R01)'],
            ['Cột chi tiết (xanh)', 'MÃ NHÓM · BẬC · ĐỊNH MỨC THỜI GIAN · SP/H · KHÂU SX · CỤM · MÔ TẢ · MÁY · MŨI · KIM/CHỈ · CỮ/GÁ · ĐƠN VỊ · NGUỒN SMV'],
            ['Cột hiệu lực (xanh đậm)', 'TRẠNG THÁI · NGÀY HL · NGÀY HẾT HL · NGƯỜI LẬP · NGƯỜI DUYỆT · LÝ DO · VIDEO_URL · NOTES'],
            [],
            ['ĐỊNH MỨC THỜI GIAN', 'Đơn vị GIÂY — lưu trực tiếp làm SMV chuẩn trên Portal. Ví dụ 36 giây.'],
            ['Trùng mã', 'Cùng MÃ CÔNG ĐOẠN + PHIÊN BẢN → hệ thống CẬP NHẬT bản ghi cũ.'],
            ['TRẠNG THÁI', 'Nháp | Thử nghiệm | Đã duyệt | Ngưng sử dụng'],
            ['Nhóm chưa có', 'Nếu MÃ NHÓM chưa tồn tại, hệ thống tự tạo nhóm tạm (có cảnh báo).'],
            ['Mẹo', 'Xuất Excel thư viện hiện tại rồi chỉnh — dễ hơn điền từ file trống.'],
            ['Dòng mẫu cũ', 'Nếu file còn SEW-1001/1002 «ví dụ» hoặc NOTES «Dòng mẫu», hệ thống sẽ bỏ qua khi import.'],
        ],
        'sample_rows': [],
        'row_builder': _library_export_rows,
    },
    KIND_ROUTING: {
        'kind': KIND_ROUTING,
        'label': 'Routing mã hàng',
        'sheet': SHEET_ROUTING,
        'headers': ROUTING_HEADERS,
        'filename_export': 'Routing_Ma_Hang',
        'filename_template': 'Mau_Import_Routing_Ma_Hang',
        'step_fill': f'Điền sheet {SHEET_ROUTING}',
        'drop_hint': f'.xlsx / .xlsm — giữ đúng tên sheet {SHEET_ROUTING}',
        'required_badges': [
            ('MÃ ĐƠN HÀNG', True),
            ('MÃ HÀNG SẢN PHẨM', True),
            ('TÊN MÃ HÀNG', False),
            ('NHÓM SẢN PHẨM', False),
            ('MÃ NHÓM', False),
            ('MÃ CÔNG ĐOẠN', True),
            ('PHIÊN BẢN', False),
            ('TÊN CÔNG ĐOẠN', False),
            ('SỐ LƯỢNG', False),
            ('BẬC CÔNG ĐOẠN', False),
            ('ĐỊNH MỨC THỜI GIAN (giây)', False),
            ('ĐỊNH MỨC THEO PHIÊN BẢN', False),
            ('TỔNG ĐỊNH MỨC', False),
            ('ĐỊNH MỨC SP/H', False),
            ('% CHÊNH LỆCH ĐỊNH MỨC THỜI GIAN', False),
            ('HỆ SỐ ĐƠN GIÁ', False),
            ('TỔNG ĐƠN GIÁ', False),
            ('TRẠNG THÁI ÁP DỤNG', False),
            ('NGÀY HIỆU LỰC', False),
            ('MÃ MÁY MÓC', False),
            ('WORK_CENTER', False),
            ('NGƯỜI LẬP', False),
            ('NOTES', False),
        ],
        'guide_rows': [
            ['HƯỚNG DẪN NHẬP ROUTING MÃ HÀNG'],
            [],
            ['Bước 1', 'Đọc sheet này — không đổi tên sheet dữ liệu.'],
            ['Bước 2', f'Điền sheet {SHEET_ROUTING}: mỗi dòng = 1 công đoạn trong quy trình mã hàng.'],
            ['Bước 3', 'Điền dữ liệu thật vào sheet dữ liệu (file mẫu chỉ có tiêu đề cột, không có dòng ví dụ).'],
            ['Bước 4', 'Portal → Routing mã hàng → Import.'],
            [],
            ['Cột bắt buộc', 'MÃ ĐƠN HÀNG · MÃ HÀNG SẢN PHẨM · MÃ CÔNG ĐOẠN'],
            ['Gom routing', 'Cùng MÃ ĐƠN HÀNG → cùng một routing (các dòng công đoạn).'],
            ['ĐỊNH MỨC THỜI GIAN', 'SMV áp dụng (giây). ĐỊNH MỨC THEO PHIÊN BẢN = SMV chuẩn (giây).'],
            ['TỔNG ĐỊNH MỨC', 'Hệ thống tự tính = SỐ LƯỢNG × ĐỊNH MỨC THỜI GIAN khi lưu.'],
            ['ĐỊNH MỨC SP/H', 'Xuất tự tính = 3600 / ĐỊNH MỨC THỜI GIAN (giây). Có thể để trống khi import.'],
            ['% CHÊNH LỆCH', 'Tự tính so với định mức theo phiên bản; >15% cần giải trình trên Portal.'],
            ['Mẹo', 'Xuất Excel routing hiện tại rồi chỉnh — dễ hơn điền từ file trống.'],
            ['Dòng mẫu cũ', 'Nếu file còn STYLE-DEMO / NOTES «Dòng mẫu», hệ thống sẽ bỏ qua khi import.'],
        ],
        'sample_rows': [],
        'row_builder': _routing_export_rows,
    },
}


def ie_dataset_meta(kind: str | None) -> dict:
    return IE_DATASETS[normalize_ie_kind(kind)]


def export_ie_dataset_workbook(kind, *, template: bool = False):
    """Workbook 2 sheet: hướng dẫn + 1 sheet dữ liệu (mẫu hoặc dữ liệu thật)."""
    meta = ie_dataset_meta(kind)
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise OperationMasterImportError('Thiếu thư viện openpyxl.') from exc

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_GUIDE
    for row in meta['guide_rows']:
        ws.append(row)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 88

    ws_data = wb.create_sheet(meta['sheet'])
    rows = meta['sample_rows'] if template else meta['row_builder']()
    _write_sheet(ws_data, meta['headers'], rows)
    return wb


def export_ie_dataset_response(kind, *, template: bool = False, user=None):
    """HttpResponse Excel theo kind (mẫu hoặc xuất dữ liệu)."""
    import io
    from datetime import datetime

    from django.http import HttpResponse

    meta = ie_dataset_meta(kind)
    wb = export_ie_dataset_workbook(kind, template=template)
    buf = io.BytesIO()
    wb.save(buf)
    stamp = datetime.now().strftime('%Y%m%d' if template else '%Y%m%d_%H%M')
    prefix = meta['filename_template'] if template else meta['filename_export']
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={prefix}_{stamp}.xlsx'
    if not template:
        from san_xuat.ie_models import SxIeAuditLog
        from san_xuat.services.ie_audit import log_ie_event

        log_ie_event(
            action=SxIeAuditLog.ACTION_EXPORT,
            summary=f'Xuất Excel {meta["label"]}',
            object_type='OperationMaster',
            object_repr=f'export_excel:{meta["kind"]}',
            user=user,
        )
    return response


def export_operation_master_workbook():
    """Tạo Workbook openpyxl cấu trúc ngược với file import (5 sheet dữ liệu chính)."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise OperationMasterImportError('Thiếu thư viện openpyxl.') from exc

    wb = Workbook()

    # 05_DM_THAM_CHIEU
    ws = wb.active
    ws.title = SHEET_REF
    machines = list(SxMachine.objects.order_by('sort_order', 'code'))
    skills = ensure_skill_levels_abc()
    sources = list(SxSmvSource.objects.order_by('sort_order', 'code'))
    stages = list(SxProcessStage.objects.order_by('sort_order', 'code'))
    stitches = list(SxStitchClass.objects.order_by('sort_order', 'code'))
    wcs = list(SxWorkCenter.objects.filter(is_demo=False).order_by('code'))
    max_n = max(len(machines), len(skills), len(sources), len(stages), len(stitches), len(wcs), 1)
    headers = [
        'MACHINE_CODE', 'MACHINE_NAME', '', 'OP_STATUS', '', 'SMV_SOURCE', '',
        'SKILL_LEVEL', '', 'YES_NO', '', 'PROCESS_STAGE', '', 'SMV_BASIS', '', 'APPROVAL_STATUS', '',
        'WORK_CENTER', '', 'STITCH_CLASS',
    ]
    ws.append(headers)
    op_statuses = ['Nháp', 'Thử nghiệm', 'Đã duyệt', 'Ngưng sử dụng']
    yes_no = ['Có', 'Không']
    approvals = ['Chờ duyệt', 'Đã duyệt', 'Từ chối', 'Cần đo lại']
    for i in range(max_n):
        ws.append([
            machines[i].code if i < len(machines) else '',
            machines[i].name if i < len(machines) else '',
            '',
            op_statuses[i] if i < len(op_statuses) else '',
            '',
            sources[i].name if i < len(sources) else '',
            '',
            skills[i].name if i < len(skills) else '',
            '',
            yes_no[i] if i < len(yes_no) else '',
            '',
            stages[i].name if i < len(stages) else '',
            '',
            approvals[i] if i < len(approvals) else '',
            '',
            wcs[i].code if i < len(wcs) else '',
            '',
            stitches[i].code if i < len(stitches) else '',
        ])

    ws = wb.create_sheet(SHEET_GROUP)
    _write_sheet(ws, GROUP_HEADERS, _group_export_rows())

    ws = wb.create_sheet(SHEET_LIB)
    _write_sheet(ws, LIB_HEADERS, _library_export_rows())

    ws = wb.create_sheet(SHEET_ROUTING)
    _write_sheet(ws, ROUTING_HEADERS, _routing_export_rows())

    # 04_DU_LIEU_TIME_STUDY
    ws = wb.create_sheet(SHEET_TIMESTUDY)
    ts_rows = []
    for t in SxTimeStudy.objects.order_by('op_code', 'obs_no', 'study_id'):
        ts_rows.append([
            t.study_id,
            t.study_date.isoformat() if t.study_date else '',
            t.factory_code, t.line_code, t.shift, t.style_code, t.routing_rev,
            t.op_code, t.op_rev, t.op_name_vi, t.operator_id, t.skill_level_label,
            t.machine_code, t.method_rev, t.obs_no,
            float(t.observed_cycle_sec or 0), float(t.abnormal_sec or 0),
            float(t.net_observed_sec or 0), float(t.performance_rating or 0),
            float(t.normal_time_sec or 0), float(t.allowance_pct or 0),
            float(t.standard_time_sec or 0), float(t.calculated_smv or 0),
            float(t.current_routing_smv or 0), float(t.variance_pct or 0),
            '', t.ie_observer, t.conditions,
            _APPROVAL_LABEL.get(t.approval_status, t.approval_status), t.notes,
        ])
    _write_sheet(ws, [
        'STUDY_ID', 'STUDY_DATE', 'FACTORY_CODE', 'LINE_CODE', 'SHIFT', 'STYLE_CODE',
        'ROUTING_REV', 'OP_CODE', 'OP_REV', 'OP_NAME_VI', 'OPERATOR_ID', 'SKILL_LEVEL',
        'MACHINE_CODE', 'METHOD_REV', 'OBS_NO', 'OBSERVED_CYCLE_SEC', 'ABNORMAL_SEC',
        'NET_OBSERVED_SEC', 'PERFORMANCE_RATING', 'NORMAL_TIME_SEC', 'ALLOWANCE_PCT',
        'STANDARD_TIME_SEC', 'CALCULATED_SMV', 'CURRENT_ROUTING_SMV', 'VARIANCE_PCT',
        'VALID_SAMPLE', 'IE_OBSERVER', 'CONDITIONS', 'APPROVAL_STATUS', 'NOTES',
    ], ts_rows)

    # 07_DASHBOARD
    from datetime import date as _date_cls

    from san_xuat.services.ie_ops import build_ie_dashboard

    dash = build_ie_dashboard()
    ws = wb.create_sheet(SHEET_DASHBOARD)
    ws.append(['JUST PLAY – CẤU TRÚC DỮ LIỆU MÃ CÔNG ĐOẠN SẢN XUẤT'])
    ws.append([])
    ws.append([])
    ws.append(['Chỉ số', 'Giá trị', '', 'Chỉ số', 'Giá trị'])
    ws.append(['Số nhóm công đoạn', dash['groups'], '', 'Số mã hàng mẫu', dash['styles']])
    ws.append(['Số công đoạn chuẩn', dash['operations'], '', 'Số quan sát time study', dash['time_studies']])
    ws.append(['Số dòng routing mẫu', dash['routing_lines'], '', 'Ngày tạo file', _date_cls.today().isoformat()])
    ws.append(['Time study đã gắn FK', dash['time_studies_linked'], '', 'OP chưa duyệt', dash['pending_ops']])
    ws.append(['Dòng lệch >15%', dash['high_var_count'], '', 'Routing chưa duyệt', dash['pending_routings']])
    ws.append([])
    ws.append([])
    ws.append([
        'STYLE_CODE', 'PRODUCT_FAMILY', 'OPERATION_COUNT', 'TOTAL_ROUTING_SMV',
        'SEWING_SMV', 'AVG_TARGET_EFFICIENCY_PCT', 'ROUTING_ID', 'APPROVAL_STATUS',
    ])
    for row in dash['style_rows']:
        ws.append([
            row['style_code'],
            row['product_family'],
            row['operation_count'],
            float(row['total_smv'] or 0),
            float(row['sewing_smv'] or 0),
            float(row['avg_target_efficiency'] or 0),
            row['routing_id'],
            row['approval_status'],
        ])
    ws.append([])
    ws.append(['CẢNH BÁO LỆCH SMV > 15%'])
    ws.append(['STYLE', 'OP_CODE', 'SEQ', 'VARIANCE_PCT', 'EXPLANATION'])
    for line in dash['high_var_lines']:
        ws.append([
            line.routing.style_code,
            line.op_code,
            line.seq_no,
            float(line.smv_variance_pct or 0),
            line.variance_explanation or '',
        ])

    return wb


def export_operation_library_template_workbook():
    """Tương thích cũ → file mẫu thư viện (2 sheet)."""
    return export_ie_dataset_workbook(KIND_LIBRARY, template=True)


def export_operation_library_template_response():
    """Tương thích cũ → HTTP file mẫu thư viện."""
    return export_ie_dataset_response(KIND_LIBRARY, template=True)


def export_operation_master_response(*, user=None):
    """HttpResponse file Excel master data mã công đoạn."""
    import io
    from datetime import datetime

    from django.http import HttpResponse

    from san_xuat.ie_models import SxIeAuditLog
    from san_xuat.services.ie_audit import log_ie_event

    wb = export_operation_master_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f'attachment; filename=Just_Play_Master_Data_Ma_Cong_Doan_{stamp}.xlsx'
    )
    log_ie_event(
        action=SxIeAuditLog.ACTION_EXPORT,
        summary='Xuất Excel master data công đoạn (gồm 07_DASHBOARD)',
        object_type='OperationMaster',
        object_repr='export_excel',
        user=user,
    )
    return response
